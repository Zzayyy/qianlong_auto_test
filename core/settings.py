# -*- coding: utf-8 -*-
"""交易系统设置的统一状态、单模块结果和批次报告。"""

from __future__ import annotations

from collections import Counter
from datetime import datetime
import json
import os
from pathlib import Path
import re
import secrets
from typing import Any, Callable, Dict, Iterable, List, Mapping, Optional


STATUS_PASS = "通过"
STATUS_DIFFERENCE = "差异"
STATUS_DISABLED = "未启用"
STATUS_UNVERIFIED = "未验证"
STATUS_ADDED = "新增"
STATUS_CONFLICT = "冲突"
STATUS_EXECUTION_FAILED = "执行失败"
STATUS_NOT_APPLICABLE = "不适用"

CANONICAL_STATUSES = frozenset({
    STATUS_PASS,
    STATUS_DIFFERENCE,
    STATUS_DISABLED,
    STATUS_UNVERIFIED,
    STATUS_ADDED,
    STATUS_CONFLICT,
    STATUS_EXECUTION_FAILED,
    STATUS_NOT_APPLICABLE,
})

DIFFERENCE_STATUSES = frozenset({
    STATUS_DIFFERENCE,
    STATUS_ADDED,
    STATUS_CONFLICT,
})

REVIEW_STATUSES = frozenset({STATUS_UNVERIFIED})

_STATUS_ALIASES: Mapping[str, str] = {
    "✓": STATUS_PASS,
    "一致": STATUS_PASS,
    "pass": STATUS_PASS,
    "passed": STATUS_PASS,
    "✗": STATUS_DIFFERENCE,
    "✗ 差异": STATUS_DIFFERENCE,
    "不一致": STATUS_DIFFERENCE,
    "diff": STATUS_DIFFERENCE,
    "different": STATUS_DIFFERENCE,
    "○ 未启用": STATUS_DISABLED,
    "未生效": STATUS_DISABLED,
    "disabled": STATUS_DISABLED,
    "unverified": STATUS_UNVERIFIED,
    "added": STATUS_ADDED,
    "conflict": STATUS_CONFLICT,
    "失败": STATUS_EXECUTION_FAILED,
    "异常": STATUS_EXECUTION_FAILED,
    "error": STATUS_EXECUTION_FAILED,
    "failed": STATUS_EXECUTION_FAILED,
    "跳过": STATUS_NOT_APPLICABLE,
    "不支持": STATUS_NOT_APPLICABLE,
    "skip": STATUS_NOT_APPLICABLE,
    "skipped": STATUS_NOT_APPLICABLE,
}


def normalize_status(status: object) -> str:
    """将历史显示值或英文状态转换为统一中文状态。

    未知状态会直接报错，避免汇总时悄悄漏掉新增状态。
    """
    value = str(status).strip()
    if value in CANONICAL_STATUSES:
        return value
    normalized = _STATUS_ALIASES.get(value)
    if normalized is None:
        normalized = _STATUS_ALIASES.get(value.lower())
    if normalized is None:
        raise ValueError(f"未知的交易系统设置状态: {status!r}")
    return normalized


def is_difference_status(status: object) -> bool:
    """返回该状态是否应计入总差异。"""
    return normalize_status(status) in DIFFERENCE_STATUSES


def count_statuses(rows: Iterable[Mapping[str, object]]) -> Counter:
    """按统一的 ``状态`` 字段统计结果行。"""
    return Counter(normalize_status(row["状态"]) for row in rows)


# ==================== 单模块检查结果 ====================


STATUS_ORDER = (
    STATUS_PASS,
    STATUS_DIFFERENCE,
    STATUS_ADDED,
    STATUS_CONFLICT,
    STATUS_UNVERIFIED,
    STATUS_DISABLED,
    STATUS_EXECUTION_FAILED,
    STATUS_NOT_APPLICABLE,
)


class SettingsTestResult:
    """统一的设置检查结果模型。

    所有结果行固定包含 ``名称/期望值/实际值/状态/说明`` 五个字段。
    ``to_file`` 在写入文本报告的同时生成同名 JSON，供批次总报告读取。
    """

    def __init__(
        self,
        panel_name: str,
        *,
        normalizer: Optional[Callable[[Any], Any]] = None,
    ):
        self.panel_name = panel_name
        self.normalizer = normalizer
        self.results: List[Dict[str, Any]] = []
        self.observations: List[Dict[str, Any]] = []

    @property
    def differences(self) -> List[Dict[str, Any]]:
        return [row for row in self.results if row["状态"] in DIFFERENCE_STATUSES]

    @property
    def unverified(self) -> List[Dict[str, Any]]:
        return [row for row in self.results if row["状态"] == STATUS_UNVERIFIED]

    @property
    def not_enabled(self) -> int:
        return sum(row["状态"] == STATUS_DISABLED for row in self.results)

    def _values_match(self, actual_value: Any, expected_value: Any) -> bool:
        if self.normalizer is None:
            return actual_value == expected_value
        return self.normalizer(actual_value) == self.normalizer(expected_value)

    def add_result(
        self,
        name: str,
        actual_value: Any,
        expected_value: Any,
        detail: str = "",
    ):
        matched = self._values_match(actual_value, expected_value)
        self.add_status(
            name,
            actual_value,
            expected_value,
            STATUS_PASS if matched else STATUS_DIFFERENCE,
            detail,
        )

    def add_status(
        self,
        name: str,
        actual_value: Any,
        expected_value: Any,
        status: str,
        detail: str = "",
    ):
        self.results.append({
            "名称": name,
            "期望值": expected_value,
            "实际值": actual_value,
            "状态": normalize_status(status),
            "说明": detail,
        })

    def add_not_enabled(self, name: str, detail: str = "(未启用)"):
        self.add_status(name, detail, "—", STATUS_DISABLED)

    def add_unverified(
        self,
        name: str,
        expected_value: Any,
        detail: str,
        actual_value: Any = "(无法确认)",
    ):
        self.add_status(
            name, actual_value, expected_value, STATUS_UNVERIFIED, detail
        )

    def add_observation(self, name: str, value: Any, detail: str = ""):
        self.observations.append({
            "名称": name,
            "采集值": value,
            "说明": detail,
        })

    def summary(self) -> Dict[str, int]:
        counts = count_statuses(self.results)
        summary = {"总项目数": len(self.results)}
        summary.update({status: counts[status] for status in STATUS_ORDER})
        summary["差异合计"] = sum(counts[status] for status in DIFFERENCE_STATUSES)
        summary["采集项"] = len(self.observations)
        return summary

    def print_summary(self):
        summary = self.summary()
        print(f"\n{'=' * 60}")
        print("测试结果汇总")
        print(f"{'=' * 60}")
        self._write_summary(print, summary)
        self._write_rows(print)

    def _write_summary(self, writer, summary: Dict[str, int]):
        writer(f"总项目数: {summary['总项目数']}")
        for status in STATUS_ORDER:
            writer(f"{status}: {summary[status]}")
        writer(f"差异合计: {summary['差异合计']}")
        writer(f"采集项: {summary['采集项']}")

    def _write_rows(self, writer):
        if self.results:
            writer("")
        for row in self.results:
            writer(f"[{row['状态']}] {row['名称']}")
            writer(f"  期望值: {row['期望值']}")
            writer(f"  实际值: {row['实际值']}")
            if row["说明"]:
                writer(f"  说明: {row['说明']}")
        if self.observations:
            writer("")
            writer("采集项（不计入差异）:")
            for row in self.observations:
                writer(f"  - {row['名称']}: {row['采集值']}")
                if row["说明"]:
                    writer(f"    说明: {row['说明']}")

    def to_payload(self, report_path: str = "") -> Dict[str, Any]:
        generated_at = datetime.now().isoformat(timespec="seconds")
        return {
            "schema_version": 1,
            "run_id": os.environ.get("GUI_SETTINGS_RUN_ID", ""),
            "module": self.panel_name,
            "client_id": os.environ.get("GUI_CLIENT_ID", ""),
            "execution_status": STATUS_PASS,
            "generated_at": generated_at,
            "report_path": os.path.abspath(report_path) if report_path else "",
            "summary": self.summary(),
            "items": [
                {
                    "name": row["名称"],
                    "expected": row["期望值"],
                    "actual": row["实际值"],
                    "status": row["状态"],
                    "detail": row["说明"],
                }
                for row in self.results
            ],
            "observations": [
                {
                    "name": row["名称"],
                    "value": row["采集值"],
                    "detail": row["说明"],
                }
                for row in self.observations
            ],
        }

    def to_file(self, filepath: str):
        run_id = os.environ.get("GUI_SETTINGS_RUN_ID", "").strip()
        run_dir = os.environ.get("GUI_SETTINGS_RUN_DIR", "").strip()
        if run_id and run_dir:
            safe_panel_name = re.sub(r'[<>:"/\\|?*]+', "_", self.panel_name).strip()
            filepath = os.path.join(run_dir, f"{safe_panel_name}.txt")

        output_dir = os.path.dirname(os.path.abspath(filepath))
        os.makedirs(output_dir, exist_ok=True)
        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        summary = self.summary()
        text_temp = f"{filepath}.tmp"
        with open(text_temp, "w", encoding="utf-8") as report:
            report.write(f"{self.panel_name}测试报告\n")
            report.write(f"生成时间: {generated_at}\n")
            report.write(f"{'=' * 60}\n\n")

            def write_line(value: str):
                report.write(f"{value}\n")

            self._write_summary(write_line, summary)
            self._write_rows(write_line)
        os.replace(text_temp, filepath)

        json_path = str(Path(filepath).with_suffix(".json"))
        json_temp = f"{json_path}.tmp"
        with open(json_temp, "w", encoding="utf-8") as structured:
            json.dump(
                self.to_payload(filepath),
                structured,
                ensure_ascii=False,
                indent=2,
                default=_json_default,
            )
            structured.write("\n")
        os.replace(json_temp, json_path)

        print(f"[OK] 测试报告已保存: {filepath}")
        print(f"[OK] 结构化结果已保存: {json_path}")
        return json_path


def _json_default(value: Any):
    """兼容 numpy 标量、Path 以及少数控件包装值。"""
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    if isinstance(value, Path):
        return str(value)
    return str(value)


# ==================== 批次汇总报告 ====================


OVERALL_PASS = "通过"
OVERALL_REVIEW = "需人工复核"
OVERALL_FAIL = "不通过"
OVERALL_RUNNING = "运行中"

BATCH_RUNNING = "运行中"
BATCH_COMPLETED = "已完成"
BATCH_STOPPED = "已停止"

BATCH_SUMMARY_JSON = "批次汇总.json"
TOTAL_REPORT_TXT = "总差异报告.txt"
TOTAL_REPORT_XLSX = "总差异报告.xlsx"
DEFAULT_BATCH_LIMIT = 20


def create_run_id(now: datetime = None, code: str = "") -> str:
    """生成可排序的“时间_专属代码”批次号。"""
    current = now or datetime.now()
    unique_code = (code or secrets.token_hex(3)).upper()
    return f"{current.strftime('%Y%m%d_%H%M%S')}_{unique_code}"


def module_name_from_script(script_name: str) -> str:
    """去掉 GUI 脚本名前的编号。"""
    return re.sub(r"^\s*\d+\.\s*", "", script_name or "").strip()


def discover_batches(
    output_dir: str,
    limit: int = DEFAULT_BATCH_LIMIT,
) -> List[Dict[str, Any]]:
    """读取最新的设置检查批次。

    先使用目录项和 ``批次汇总.json`` 的修改时间排序，再只解析需要显示
    的 JSON，避免历史批次较多时逐个加载全部汇总内容。
    """
    batch_root = Path(output_dir) / "批次"
    if not batch_root.is_dir():
        return []
    candidates = []
    try:
        entries = list(os.scandir(batch_root))
    except OSError:
        return []
    for entry in entries:
        try:
            if not entry.is_dir(follow_symlinks=False):
                continue
            summary_path = Path(entry.path) / BATCH_SUMMARY_JSON
            stat = summary_path.stat()
        except OSError:
            continue
        candidates.append((stat.st_mtime_ns, entry.name, summary_path))
    candidates.sort(key=lambda row: (row[0], row[1]), reverse=True)

    batches = []
    max_count = max(0, int(limit)) if limit is not None else None
    for _, _, summary_path in candidates:
        if max_count is not None and len(batches) >= max_count:
            break
        try:
            payload = json.loads(summary_path.read_text(encoding="utf-8-sig"))
        except Exception:
            continue
        child = summary_path.parent
        payload["_batch_dir"] = str(child.resolve())
        payload["_summary_path"] = str(summary_path.resolve())
        batches.append(payload)
    return batches


def load_module_results(batch_dir: str) -> Dict[str, Dict[str, Any]]:
    """读取一个批次目录中的模块 JSON，以模块名为键。"""
    root = Path(batch_dir)
    loaded: Dict[str, Dict[str, Any]] = {}
    if not root.is_dir():
        return loaded
    for path in root.rglob("*.json"):
        if path.name == BATCH_SUMMARY_JSON:
            continue
        try:
            payload = json.loads(path.read_text(encoding="utf-8-sig"))
        except Exception:
            continue
        module = str(payload.get("module", "")).strip()
        if not module or not isinstance(payload.get("items"), list):
            continue
        payload["_json_path"] = str(path.resolve())
        previous = loaded.get(module)
        if previous is None or str(payload.get("generated_at", "")) >= str(
            previous.get("generated_at", "")
        ):
            loaded[module] = payload
    return loaded


def build_batch_summary(
    *,
    run_id: str,
    batch_dir: str,
    client_id: str,
    task_records: Iterable[Mapping[str, Any]],
    stopped: bool = False,
    source: str = "",
    batch_status: str = BATCH_COMPLETED,
) -> Dict[str, Any]:
    """将任务退出状态和各模块 JSON 合并成批次汇总。"""
    loaded = load_module_results(batch_dir)
    modules: List[Dict[str, Any]] = []
    problems: List[Dict[str, Any]] = []

    totals = {
        "模块数": 0,
        STATUS_PASS: 0,
        STATUS_DIFFERENCE: 0,
        STATUS_ADDED: 0,
        STATUS_CONFLICT: 0,
        STATUS_UNVERIFIED: 0,
        STATUS_DISABLED: 0,
        STATUS_EXECUTION_FAILED: 0,
        STATUS_NOT_APPLICABLE: 0,
        "差异合计": 0,
        "采集项": 0,
    }

    for task in task_records:
        if task.get("category") != "交易系统设置":
            continue
        module = module_name_from_script(str(task.get("script_name", "")))
        task_status = str(task.get("status", ""))
        payload = loaded.get(module)
        return_code = task.get("return_code")
        elapsed = float(task.get("elapsed", 0.0) or 0.0)
        task_error = str(task.get("error", "") or "")

        execution_ok = task_status == "成功" and return_code in (None, 0)
        if execution_ok and payload is None:
            execution_ok = False
            task_error = task_error or "脚本执行成功，但没有生成结构化结果 JSON"

        if execution_ok:
            raw_summary = payload.get("summary", {})
            module_summary = {
                status: int(raw_summary.get(status, 0) or 0)
                for status in (
                    STATUS_PASS,
                    STATUS_DIFFERENCE,
                    STATUS_ADDED,
                    STATUS_CONFLICT,
                    STATUS_UNVERIFIED,
                    STATUS_DISABLED,
                    STATUS_NOT_APPLICABLE,
                )
            }
            difference_total = int(
                raw_summary.get(
                    "差异合计",
                    sum(module_summary.get(status, 0) for status in DIFFERENCE_STATUSES),
                )
                or 0
            )
            if difference_total:
                conclusion = OVERALL_FAIL
            elif module_summary[STATUS_UNVERIFIED]:
                conclusion = OVERALL_REVIEW
            else:
                conclusion = OVERALL_PASS
            module_row = {
                "module": module,
                "execution_status": STATUS_PASS,
                "conclusion": conclusion,
                "total": int(raw_summary.get("总项目数", 0) or 0),
                "passed": module_summary[STATUS_PASS],
                "differences": difference_total,
                "unverified": module_summary[STATUS_UNVERIFIED],
                "disabled": module_summary[STATUS_DISABLED],
                "elapsed": elapsed,
                "report_path": str(payload.get("report_path", "") or ""),
                "json_path": str(payload.get("_json_path", "") or ""),
                "detail": "",
            }
            for key in (
                STATUS_PASS,
                STATUS_DIFFERENCE,
                STATUS_ADDED,
                STATUS_CONFLICT,
                STATUS_UNVERIFIED,
                STATUS_DISABLED,
                STATUS_NOT_APPLICABLE,
            ):
                totals[key] += module_summary[key]
            totals["差异合计"] += difference_total
            totals["采集项"] += int(raw_summary.get("采集项", 0) or 0)

            for item in payload.get("items", []):
                status = str(item.get("status", ""))
                if status not in DIFFERENCE_STATUSES and status != STATUS_UNVERIFIED:
                    continue
                problems.append({
                    "status": status,
                    "module": module,
                    "name": item.get("name", ""),
                    "expected": item.get("expected", ""),
                    "actual": item.get("actual", ""),
                    "detail": item.get("detail", ""),
                    "report_path": module_row["report_path"],
                })
        else:
            detail = task_error
            if not detail:
                if task_status == "已停止":
                    detail = "用户停止了本批次"
                elif return_code not in (None, 0):
                    detail = f"脚本退出码: {return_code}"
                else:
                    detail = f"任务状态: {task_status or '未知'}"
            module_row = {
                "module": module,
                "execution_status": STATUS_EXECUTION_FAILED,
                "conclusion": OVERALL_FAIL,
                "total": 0,
                "passed": 0,
                "differences": 0,
                "unverified": 0,
                "disabled": 0,
                "elapsed": elapsed,
                "report_path": "",
                "json_path": "",
                "detail": detail,
            }
            totals[STATUS_EXECUTION_FAILED] += 1
            problems.append({
                "status": STATUS_EXECUTION_FAILED,
                "module": module,
                "name": "模块执行",
                "expected": "执行成功并生成结果",
                "actual": task_status or "未知",
                "detail": detail,
                "report_path": "",
            })

        modules.append(module_row)

    totals["模块数"] = len(modules)
    if batch_status == BATCH_RUNNING:
        overall = OVERALL_RUNNING
    elif stopped or totals[STATUS_EXECUTION_FAILED] or totals["差异合计"]:
        overall = OVERALL_FAIL
    elif totals[STATUS_UNVERIFIED]:
        overall = OVERALL_REVIEW
    else:
        overall = OVERALL_PASS

    return {
        "schema_version": 1,
        "run_id": run_id,
        "client_id": client_id,
        "source": source,
        "batch_status": batch_status,
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "overall_status": overall,
        "stopped": bool(stopped),
        "batch_dir": str(Path(batch_dir).resolve()),
        "totals": totals,
        "modules": modules,
        "problems": problems,
    }


def write_batch_reports(summary: Dict[str, Any]) -> Dict[str, Any]:
    """写入批次汇总 JSON、总差异 TXT 和 Excel。"""
    batch_dir = Path(summary["batch_dir"])
    batch_dir.mkdir(parents=True, exist_ok=True)
    txt_path = batch_dir / TOTAL_REPORT_TXT
    xlsx_path = batch_dir / TOTAL_REPORT_XLSX
    json_path = batch_dir / BATCH_SUMMARY_JSON

    summary["txt_path"] = str(txt_path.resolve())
    summary["xlsx_path"] = str(xlsx_path.resolve())
    summary["summary_path"] = str(json_path.resolve())

    txt_temp = batch_dir / f".{TOTAL_REPORT_TXT}.tmp"
    xlsx_temp = batch_dir / f".{TOTAL_REPORT_XLSX}.tmp"
    json_temp = batch_dir / f".{BATCH_SUMMARY_JSON}.tmp"
    _write_text_report(summary, txt_temp)
    _write_excel_report(summary, xlsx_temp)
    json_temp.write_text(
        json.dumps(summary, ensure_ascii=False, indent=2, default=str) + "\n",
        encoding="utf-8",
    )
    os.replace(txt_temp, txt_path)
    os.replace(xlsx_temp, xlsx_path)
    os.replace(json_temp, json_path)
    return summary


def generate_batch_reports(**kwargs) -> Dict[str, Any]:
    return write_batch_reports(build_batch_summary(**kwargs))


def _write_text_report(summary: Mapping[str, Any], path: Path):
    totals = summary["totals"]
    lines = [
        "交易系统设置总差异报告",
        f"批次号: {summary['run_id']}",
        f"客户端: {summary.get('client_id', '')}",
        f"运行来源: {summary.get('source', '')}",
        f"批次状态: {summary.get('batch_status', '')}",
        f"生成时间: {summary['generated_at']}",
        f"总体结论: {summary['overall_status']}",
        "=" * 72,
        "",
        f"执行模块: {totals['模块数']}",
        f"通过项: {totals[STATUS_PASS]}",
        f"差异: {totals[STATUS_DIFFERENCE]}",
        f"新增: {totals[STATUS_ADDED]}",
        f"冲突: {totals[STATUS_CONFLICT]}",
        f"未验证: {totals[STATUS_UNVERIFIED]}",
        f"未启用: {totals[STATUS_DISABLED]}",
        f"执行失败模块: {totals[STATUS_EXECUTION_FAILED]}",
        f"差异合计: {totals['差异合计']}",
        "",
        "模块汇总:",
    ]
    for module in summary["modules"]:
        lines.append(
            "  - {module}: {conclusion}；检查项={total}，通过={passed}，"
            "差异={differences}，未验证={unverified}，未启用={disabled}，"
            "耗时={elapsed:.1f}s".format(**module)
        )
        if module["detail"]:
            lines.append(f"    说明: {module['detail']}")

    lines.extend(["", "问题明细:"])
    if not summary["problems"]:
        lines.append("  无")
    for problem in summary["problems"]:
        lines.extend([
            f"[{problem['status']}] {problem['module']} / {problem['name']}",
            f"  期望值: {problem['expected']}",
            f"  实际值: {problem['actual']}",
        ])
        if problem["detail"]:
            lines.append(f"  说明: {problem['detail']}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _write_excel_report(summary: Mapping[str, Any], path: Path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    workbook = Workbook()
    overview = workbook.active
    overview.title = "模块汇总"
    problem_sheet = workbook.create_sheet("问题明细")

    overview.append(["批次号", summary["run_id"]])
    overview.append(["客户端", summary.get("client_id", "")])
    overview.append(["运行来源", summary.get("source", "")])
    overview.append(["批次状态", summary.get("batch_status", "")])
    overview.append(["生成时间", summary["generated_at"]])
    overview.append(["总体结论", summary["overall_status"]])
    overview.append([])
    overview.append([
        "模块", "执行状态", "结论", "检查项", "通过", "差异",
        "未验证", "未启用", "耗时(秒)", "说明", "单项报告",
    ])
    for module in summary["modules"]:
        overview.append([
            module["module"],
            module["execution_status"],
            module["conclusion"],
            module["total"],
            module["passed"],
            module["differences"],
            module["unverified"],
            module["disabled"],
            round(module["elapsed"], 1),
            module["detail"],
            module["report_path"],
        ])

    problem_sheet.append([
        "状态", "模块", "检查项", "期望值", "实际值", "说明", "单项报告",
    ])
    for problem in summary["problems"]:
        problem_sheet.append([
            problem["status"],
            problem["module"],
            problem["name"],
            str(problem["expected"]),
            str(problem["actual"]),
            problem["detail"],
            problem["report_path"],
        ])

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    fail_fill = PatternFill("solid", fgColor="FDE9E7")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    pass_fill = PatternFill("solid", fgColor="E2F0D9")

    for sheet, header_row in ((overview, 8), (problem_sheet, 1)):
        for cell in sheet[header_row]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.freeze_panes = f"A{header_row + 1}"
        sheet.auto_filter.ref = sheet.dimensions
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        for column in range(1, sheet.max_column + 1):
            max_length = max(
                len(str(sheet.cell(row=row, column=column).value or ""))
                for row in range(1, sheet.max_row + 1)
            )
            sheet.column_dimensions[get_column_letter(column)].width = min(
                max(max_length + 2, 10), 42
            )

    overview["B6"].font = Font(bold=True)
    overview["B6"].fill = {
        OVERALL_PASS: pass_fill,
        OVERALL_REVIEW: review_fill,
        OVERALL_FAIL: fail_fill,
        OVERALL_RUNNING: header_fill,
    }[summary["overall_status"]]
    for row in range(2, problem_sheet.max_row + 1):
        status = problem_sheet.cell(row=row, column=1).value
        fill = review_fill if status == STATUS_UNVERIFIED else fail_fill
        for cell in problem_sheet[row]:
            cell.fill = fill

    workbook.save(path)
