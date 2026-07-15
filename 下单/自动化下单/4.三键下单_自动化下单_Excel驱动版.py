# -*- coding: utf-8 -*-
"""
钱龙期权交易 - 三键下单面板自动化(Excel文件驱动版)
==============================================
功能流程:
    1. 从 Excel 文件读取下单配置(菜单/合约代码/报价方式/委托数量/动作/备兑/自动/FOK)
    2. 查找"钱龙"主窗口
    3. 激活窗口并切换到"三键下单"面板
    4. 按 Excel 每行配置执行下单:
       - 录入合约代码
       - 选择报价方式
       - 设置委托数量
       - 勾选备兑/自动/FOK
       - 点击对应动作按钮(买开/卖开/平仓)
    5. 确认弹窗

Excel 字段说明:
    菜单     : 面板名称(如"三键下单")
    合约代码 : 期权合约代码
    报价方式 : 对手价/挂盘价/涨停价/跌停价/限价/超价/市价转限/市价FAK/市价FOK
    委托数量 : 纯数字(如"5") / 百分比(如"30%") / "FOK"
    动作     : 买开 / 卖开 / 平仓
    备兑     : True=勾选, False=不勾选
    自动     : True=勾选, False=不勾选
    FOK      : True=勾选, False=不勾选

运行环境:
    pip install pywinauto==0.6.8 openpyxl

使用方法:
    1. 打开钱龙旗舰版,登录交易账号
    2. 准备 Excel 文件,填入下单配置(菜单列填写"三键下单")
    3. 修改下方 EXCEL_PATH 为你的 Excel 文件路径
    4. 运行本脚本(将鼠标快速移到屏幕左上角可紧急停止)
    5. 倒计时内切换到钱龙窗口
"""

from pywinauto import Application, findwindows
from core.window import switch_panel
import os
import time
import sys
import ctypes
import openpyxl


# ====================== 可配置参数 ======================
# ---- Excel 路径:优先读取 GUI 传入的环境变量,否则使用默认路径 ----
_gui_xlsx = os.environ.get("GUI_XLSX_FILE", "").strip()
if _gui_xlsx and os.path.exists(_gui_xlsx):
    EXCEL_PATH = _gui_xlsx
else:
    EXCEL_PATH = r"C:\Users\Administrator\Desktop\新建 XLSX 工作表.xlsx"
WINDOW_KEY = "钱龙模拟期权宝"        # 窗口标题关键字
USE_ENTER_CONFIRM = True  # True=按回车(推荐), False=鼠标点"确定"
COUNTDOWN = 3              # 操作前倒计时秒数
INTERVAL = 1.0             # 每两次下单之间的间隔(秒)
# ========================================================


# 报价方式映射: Excel中填写值 -> 实际下拉框显示值
QUOTE_TYPE_MAP = {
    "市价FOK": "市价F0K",   # 实际控件是 F0K(数字0),不是 FOK(字母O)
}

# 报价方式可选项(以软件实际显示为准)
# QUOTE_TYPES = [
#     "对手价",
#     "挂盘价",
#     "涨停价",
#     "跌停价",
#     "限价",
#     "超价",
#     "市价转限",
#     "市价FAK",
#     "市价F0K",
# ]

QUOTE_AUTO_ID = "18083"       # 报价方式输入框 auto_id
QTY_AUTO_ID   = "306"         # 委托数量输入框 auto_id

# 下单按钮映射: 动作名称 -> 控件 auto_id
ACTION_BUTTONS = {
    "买开": "18066",   # 实际 title: 买多
    "卖开": "18065",   # 实际 title: 卖空
    "平仓": "18070",   # 实际 title: 平仓
}

# 复选框 auto_id 映射
CHECKBOX_AUTO_IDS = {
    "备兑": "5025",
    "FOK": "5027",
    "自动": "5028",
}


def read_excel(path: str) -> list:
    """读取 Excel 文件,返回配置列表(只读取菜单列为"三键下单"的行)。"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    # 标准化列名(去除空格)
    headers = [str(h).strip() if h else "" for h in headers]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        item = dict(zip(headers, row))
        # 只读取菜单列为"三键下单"的行
        menu_value = str(item.get("菜单", "")).strip()
        if menu_value != "三键下单":
            continue
        rows.append(item)

    wb.close()
    print(f"[OK] 读取 Excel: {path}, 共 {len(rows)} 条配置(菜单列过滤: 三键下单)")
    return rows


def find_window(keyword: str):
    """根据关键字查找窗口,返回第一个匹配的句柄。"""
    elements = findwindows.find_elements(title_re=f".*{keyword}.*")
    if not elements:
        raise RuntimeError(f"未找到包含'{keyword}'的窗口,请确认软件已启动")
    return elements[0].handle


def activate_window(hwnd: int):
    """连接窗口并置前。"""
    app = Application(backend="uia").connect(handle=hwnd)
    win = app.window(handle=hwnd)
    win.set_focus()
    return win


def fill_contract_code(win, code: str):
    """双击合约代码输入框,清空后写入新代码。"""
    edit = win.child_window(auto_id="18005", control_type="Edit")
    edit.wait("ready", timeout=5)
    edit.double_click_input()
    time.sleep(0.2)
    edit.type_keys("^a", with_spaces=False)
    edit.set_edit_text(str(code))
    print(f"[OK] 已录入合约代码: {code}")


def click_lock(win):
    """点击"锁定"按钮,触发合约信息刷新。"""
    btn = win.child_window(title="锁定", control_type="Button")
    btn.wait("visible", timeout=5)
    btn.click_input()
    print("[OK] 已点击'锁定'按钮")


def confirm_auto_dialog(main_win=None, timeout: float = 8):
    """确认'自动净仓'弹窗(回车确认)。

    通过进程 ID 过滤,避免把回车发到其它程序窗口
    (例如打开着的 "平仓.xlsx - Excel")。
    """
    main_hwnd = None
    main_pid = None
    if main_win is not None:
        try:
            main_hwnd = main_win.handle
        except Exception:
            main_hwnd = None
        try:
            main_pid = main_win.process_id()
        except Exception:
            if main_hwnd is not None:
                pid = ctypes.c_ulong(0)
                ctypes.windll.user32.GetWindowThreadProcessId(
                    main_hwnd, ctypes.byref(pid))
                main_pid = pid.value or None

    end = time.time() + timeout
    while time.time() < end:
        try:
            elems = findwindows.find_elements(top_level_only=True)
        except Exception:
            elems = []
        for elem in elems:
            # 进程过滤: 只处理与主窗口同进程的窗口
            if main_pid is not None:
                try:
                    if elem.process_id != main_pid:
                        continue
                except Exception:
                    continue
            hwnd = elem.handle
            if main_hwnd is not None and hwnd == main_hwnd:
                continue
            try:
                title = (elem.name or "").strip()
            except Exception:
                title = ""
            if "自动净仓" not in title:
                continue
            try:
                dlg_app = Application(backend="uia").connect(handle=hwnd, timeout=0.5)
                dlg = dlg_app.window(handle=hwnd)
                dlg.set_focus()
                time.sleep(0.1)
                dlg.type_keys("{ENTER}", with_spaces=False)
                print(f"[OK] 自动净仓弹窗已确认 (hwnd={hwnd}, title='{title}', pid={main_pid})")
                return True
            except Exception:
                continue
        time.sleep(0.15)
    print(f"[WARN] 等待'自动净仓'弹窗超时({timeout}s)")


def get_checkbox(win, name: str, cache: dict = None):
    """获取复选框控件,支持缓存句柄(同一会话内复用,避免重复查找)。"""
    auto_id = CHECKBOX_AUTO_IDS[name]
    if cache is not None:
        cached = cache.get(name)
        if cached is not None:
            try:
                cached.is_enabled()  # 探测句柄是否仍有效
                return cached
            except Exception:
                cache.pop(name, None)
    try:
        cb = win.child_window(auto_id=auto_id, control_type="CheckBox")
        cb.wait("ready", timeout=3)
    except Exception:
        # 控件可能因其它选项(如勾选"自动"会让"备兑"置灰)而暂时改变类型/不可用,
        # 此时仅按 auto_id 再查一次,使其能走"灰色跳过"的统一分支
        try:
            cb = win.child_window(auto_id=auto_id)
            cb.wait("ready", timeout=3)
        except Exception:
            return None
    if cache is not None:
        cache[name] = cb
    return cb


def set_checkbox(win, name: str, enable: bool, cache: dict = None):
    """设置复选框到目标状态(直接 toggle,无需移动鼠标,速度更快)。

    优化点:
      1. 使用 UIA TogglePattern(cb.toggle())代替 click_input(),避免鼠标移动/点击
      2. 仅在当前状态与目标不一致时才操作,无冗余动作
      3. toggle 失效时自动回退到 click_input(),保证可靠性
    """
    if name not in CHECKBOX_AUTO_IDS:
        raise ValueError(f"未知复选框: {name!r}")

    cb = get_checkbox(win, name, cache)
    if cb is None:
        print(f"[WARN] 复选框 {name} 未找到(可能受其它选项影响被置灰/隐藏),跳过")
        return

    # 检查是否为灰色(不可用)状态
    if not cb.is_enabled():
        if enable:
            print(f"[WARN] 复选框 {name} 为灰色(不可用),无法勾选(可能'自动'已勾选),跳过")
        else:
            print(f"[--] 复选框 {name} 为灰色(不可用),跳过")
        return

    is_checked = cb.get_toggle_state() == 1
    if enable == is_checked:
        print(f"[--] 复选框 {name} 已是{'勾选' if enable else '取消勾选'}状态,跳过")
        return

    # 优先用 UIA TogglePattern 快速切换(不移动鼠标)
    try:
        cb.toggle()
    except Exception:
        cb.click_input()

    # 验证是否生效,未生效则回退到真实点击
    try:
        if (cb.get_toggle_state() == 1) != enable:
            cb.click_input()
    except Exception:
        pass

    if enable:
        print(f"[OK] 勾选复选框: {name}")
        # 勾选"自动"时会弹出"自动净仓"确认弹窗,需回车确认
        if name == "自动":
            time.sleep(0.2)
            confirm_auto_dialog(main_win=win)
    else:
        print(f"[OK] 取消勾选复选框: {name}")


def set_all_checkboxes(win, enable_beidui: bool, enable_zidong: bool, enable_fok: bool, cache: dict = None):
    """统一设置备兑、自动、FOK 三个复选框(直接设定到目标状态,无需预重置)。

    设置顺序说明:必须先确定"自动"的最终状态,再设"备兑"。
    因为软件中勾选"自动"后"备兑"会置灰,若先设"备兑"再设"自动",
    上一单遗留的"自动=勾选"会让本单"备兑"一开始就是灰色而设不上去。
    先设"自动":若目标为取消,则"备兑"解冻;若目标为勾选,则"备兑"本就该保持未勾选。
    """
    set_checkbox(win, "自动", enable_zidong, cache=cache)
    set_checkbox(win, "备兑", enable_beidui, cache=cache)
    set_checkbox(win, "FOK", enable_fok, cache=cache)


def select_quote_type(win, option: str, auto_id: str = QUOTE_AUTO_ID):
    """点击"报价方式"输入框弹出下拉,再从 ListItem 中选中目标项。"""
    # 映射: Excel填写值 -> 实际控件文本
    actual_option = QUOTE_TYPE_MAP.get(option, option)

    edit = win.child_window(auto_id=auto_id, control_type="Edit")
    edit.wait("ready", timeout=5)
    edit.click_input()
    time.sleep(0.5)

    # 方案 A: 直接定位下拉项
    try:
        item = win.child_window(title=actual_option, control_type="ListItem")
        if item.exists(timeout=1):
            item.wait("visible", timeout=2)
            item.click_input()
            print(f"[OK] 报价方式: {option} -> {actual_option} (ListItem 点击)")
            return
    except Exception:
        pass

    # 方案 B: 键盘导航兜底
    edit.type_keys("^a", with_spaces=False)
    time.sleep(0.1)
    edit.type_keys("{END}", with_spaces=False)
    time.sleep(0.1)
    edit.type_keys("{ENTER}", with_spaces=False)
    print(f"[OK] 报价方式(键盘兜底): {option} -> {actual_option}")


def click_action_button(win, action: str):
    """点击下单动作按钮(买开/卖开/平仓)。"""
    if action not in ACTION_BUTTONS:
        raise ValueError(f"未知动作: {action!r},可选: {list(ACTION_BUTTONS.keys())}")

    auto_id = ACTION_BUTTONS[action]
    btn = win.child_window(auto_id=auto_id, control_type="Button")
    btn.wait("ready", timeout=5)
    btn.click_input()
    print(f"[OK] 下单动作: {action} (auto_id={auto_id})")


def press_enter_to_confirm(
    main_win=None,
    dialog_patterns: list = None,
    timeout: float = 3,
):
    """通过回车键确认弹窗。

    通过进程 ID 过滤,避免把回车发送到其它程序窗口
    (例如打开着的 "平仓.xlsx - Excel" 也会命中 "平仓" 关键字)。
    """
    if dialog_patterns is None:
        dialog_patterns = ["三键下单", "提示", "确认", "确定"]

    main_hwnd = None
    main_pid = None
    if main_win is not None:
        try:
            main_hwnd = main_win.handle
        except Exception:
            main_hwnd = None
        try:
            main_pid = main_win.process_id()
        except Exception:
            if main_hwnd is not None:
                pid = ctypes.c_ulong(0)
                ctypes.windll.user32.GetWindowThreadProcessId(
                    main_hwnd, ctypes.byref(pid))
                main_pid = pid.value or None

    # 高优先级关键字(明确是弹窗)优先于宽泛关键字(平仓/反手)
    high_priority = ("提示", "确认", "确定", "委托确认", "风险提示")

    end = time.time() + timeout
    while time.time() < end:
        try:
            elems = findwindows.find_elements(top_level_only=True)
        except Exception:
            elems = []

        high_candidates = []
        low_candidates = []
        for elem in elems:
            # 1) 进程过滤: 只处理与主窗口同进程的窗口
            if main_pid is not None:
                try:
                    if elem.process_id != main_pid:
                        continue
                except Exception:
                    continue
            hwnd = elem.handle
            # 2) 跳过主窗口本身
            if main_hwnd is not None and hwnd == main_hwnd:
                continue
            try:
                title = (elem.name or "").strip()
            except Exception:
                title = ""
            if not any(p in title for p in dialog_patterns):
                continue
            if any(p in title for p in high_priority):
                high_candidates.append((hwnd, title))
            else:
                low_candidates.append((hwnd, title))

        for hwnd, title in high_candidates + low_candidates:
            try:
                dlg_app = Application(backend="uia").connect(handle=hwnd, timeout=0.5)
                dlg = dlg_app.window(handle=hwnd)
                dlg.set_focus()
                time.sleep(0.1)
                dlg.type_keys("{ENTER}", with_spaces=False)
                print(f"[OK] 回车确认 (hwnd={hwnd}, title='{title}', pid={main_pid})")
                return True
            except Exception as e:
                print(f"[--] 弹窗(hwnd={hwnd}, title='{title}')回车失败: {e}")
                continue

        time.sleep(0.15)
    print(f"[WARN] 等待弹窗超时({timeout}s),匹配: {dialog_patterns}")
    return False


def confirm_all_dialogs(
    main_win=None,
    max_dialogs: int = 5,
    no_dialog_timeout: float = 2.0,
    per_dialog_timeout: float = 3.0,
    use_enter: bool = USE_ENTER_CONFIRM,
):
    """自动确认所有弹窗，直到一段时间内没有新弹窗出现。
    
    Args:
        main_win: 主窗口对象,用于进程过滤,避免误操作其它程序窗口
        max_dialogs: 最大弹窗数量上限（防止死循环）
        no_dialog_timeout: 等待新弹窗的超时时间（秒），超过此时间无新弹窗则认为全部处理完毕
        per_dialog_timeout: 单个弹窗的等待超时时间（秒）
        use_enter: 是否使用回车确认
    """
    count = 0
    for i in range(1, max_dialogs + 1):
        print(f"[..] 等待第 {i} 个弹窗 (超时{no_dialog_timeout}s无新弹窗则结束)...")
        ok = press_enter_to_confirm(main_win=main_win, timeout=per_dialog_timeout)
        if ok:
            count += 1
            print(f"[OK] 已确认第 {count} 个弹窗")
            time.sleep(0.4)  # 弹窗间短暂间隔
        else:
            # 超时未出现弹窗，认为全部处理完毕
            print(f"[OK] 无更多弹窗，共确认 {count} 个")
            break
    else:
        print(f"[WARN] 达到最大弹窗数量上限({max_dialogs})")


def fill_order_quantity(win, value, auto_id: str = QTY_AUTO_ID):
    """设置委托数量。"""
    edit = win.child_window(auto_id=auto_id, control_type="Edit")
    edit.wait("ready", timeout=5)

    str_value = str(value)

    # 纯数字路径
    if str_value.isdigit():
        edit.double_click_input()
        time.sleep(0.15)
        edit.type_keys("^a", with_spaces=False)
        edit.set_edit_text(str_value)
        print(f"[OK] 委托数量(直写): {str_value}")
        return

    # 百分比 / FOK 路径
    edit.click_input()
    time.sleep(0.5)

    candidates = [str_value, str_value.replace("%", ""), f"{str_value}%"]
    for title in candidates:
        try:
            item = win.child_window(title=title, control_type="Button")
            if item.exists(timeout=1):
                item.wait("visible", timeout=2)
                item.click_input()
                print(f"[OK] 委托数量(下拉): {title}")
                return
        except Exception:
            continue

    # 兜底:键盘输入
    print(f"[WARN] 下拉未匹配到 {str_value},改用键盘输入")
    edit.type_keys("^a", with_spaces=False)
    edit.set_edit_text("")
    edit.type_keys(str_value, with_spaces=False)
    edit.type_keys("{ENTER}", with_spaces=False)


def countdown(seconds: int):
    """操作前倒计时,让用户有时间切换窗口。"""
    print(f"将在 {seconds} 秒后开始,请把焦点切到钱龙软件...")
    try:
        for i in range(seconds, 0, -1):
            print(f"  {i}...", end="\r")
            time.sleep(1)
    except KeyboardInterrupt:
        raise
    print(" " * 30, end="\r")


def normalize_bool(value) -> bool:
    """将 Excel 中的布尔值或字符串转换为 Python bool。"""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        s = value.strip().lower()
        if s in ("true", "是", "1", "yes"):
            return True
        elif s in ("false", "否", "0", "no"):
            return False
    return bool(value)


def main():
    try:
        # 读取 Excel 配置
        configs = read_excel(EXCEL_PATH)
        if not configs:
            print("[错误] Excel 文件为空或无有效数据")
            sys.exit(1)

        # 从第一行获取面板名称,未指定则用默认值
        first_config = configs[0]
        tree_item = first_config.get("菜单", "三键下单")
        if not tree_item:
            tree_item = "三键下单"

        print(f"计划执行: {len(configs)} 次下单")
        print(f"面板: {tree_item}")

        countdown(COUNTDOWN)

        hwnd = find_window(WINDOW_KEY)
        print(f"[OK] 已找到窗口,句柄 = {hwnd}")

        win = activate_window(hwnd)
        switch_panel(win, tree_item)
        time.sleep(0.5)

        cb_cache = {}  # 复选框控件缓存,整个下单过程复用句柄

        for idx, cfg in enumerate(configs, 1):
            contract_code = str(cfg.get("合约代码", "")).strip()
            quote_type = str(cfg.get("报价方式", "")).strip()
            order_qty = cfg.get("委托数量", "1")
            # 小数 -> *100 + "%"
            if isinstance(order_qty, float):
                order_qty = f"{int(order_qty * 100)}%"
            else:
                order_qty = str(order_qty).strip()
            action = str(cfg.get("动作", "")).strip()
            enable_beidui = normalize_bool(cfg.get("备兑", False))
            enable_zidong = normalize_bool(cfg.get("自动", False))
            enable_fok = normalize_bool(cfg.get("FOK", False))

            print(f"\n=== [{idx}/{len(configs)}] 合约={contract_code} "
                  f"报价={quote_type} 动作={action} ===")

            if not contract_code:
                print("[WARN] 合约代码为空,跳过")
                continue

            # 录入合约代码
            fill_contract_code(win, contract_code)
            time.sleep(0.3)

            # 设置委托数量
            fill_order_quantity(win, order_qty)
            time.sleep(0.3)

            # 选择报价方式
            if quote_type:
                select_quote_type(win, quote_type)
                time.sleep(0.3)

            # 直接设置复选框到目标状态(状态比较已保证正确,无需先全部取消)
            set_all_checkboxes(win, enable_beidui, enable_zidong, enable_fok, cache=cb_cache)
            time.sleep(0.1)

            # 执行下单动作
            if action:
                click_action_button(win, action)
                time.sleep(1)

                # 自动确认所有弹窗(无论2个还是3个)
                confirm_all_dialogs(main_win=win)

            time.sleep(INTERVAL)

        print(f"\n=== 全部完成: {len(configs)} 次下单 ===")

    except KeyboardInterrupt:
        print("\n[中断] 用户主动停止")
        sys.exit(0)
    except Exception as e:
        print(f"\n[错误] {type(e).__name__}: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
