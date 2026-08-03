# -*- coding: utf-8 -*-
"""GUI自动化主窗口：界面构建、参数配置、日志展示与执行编排"""

import os
import sys
import time
import shutil
import logging
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from config import (
    get_scripts_config,
    CATEGORIES,
    MODULE_GROUPS,
    SUPER_STRATEGY_CATEGORIES,
    get_module_for_category,
    load_user_config,
    save_user_config,
    get_output_dir,
    set_output_dir,
    get_script_filename,
    IS_FROZEN,
    PROJECT_ROOT,
    get_client,
    get_client_ids,
    get_client_name,
    get_default_client_id,
    CAPTURE_STANDARD_PANELS,
    DEFAULT_SUPER_STRATEGY_UNDERLYING,
    SUPER_STRATEGY_UNDERLYINGS,
    SUPER_STRATEGY_COMBO_MARKETS,
    SUPER_STRATEGY_COMBO_STRATEGIES,
)
from engine.runner import ScriptRunner
from engine.task import Task
from engine.scheduler import TaskScheduler
from gui.widgets import ColoredLogText
from gui.task_center import TaskCenter
from gui.scheduler_view import SchedulerPanel
from gui.history import (
    HistoryManager,
    STATUS_SUCCESS,
    STATUS_FAILED,
    STATUS_ERROR,
    STATUS_STOPPED,
    STATUS_RUNNING,
)
from gui.history_panel import HistoryPanel
from gui.compare import ComparePanel
from gui.settings_report import SettingsReportPanel
from gui.shell_open import open_path
from core.settings_standard import resolve_standard_path, STANDARD_ROOT
from core.settings import (
    BATCH_COMPLETED,
    BATCH_STOPPED,
    create_run_id,
    generate_batch_reports,
)


class AutomationGUI:
    """GUI自动化主界面"""

    # 组合申报：全自动脚本只配置委托数量，查询类脚本配置导出参数
    COMBO_AUTO_SCRIPTS = ("1.组合申报_全自动", "2.拆分申报_全自动")

    # 超级策略分类下需要选择"市场/策略/组合数量"参数（但合约一/合约二不固定，
    # 由运行期按持仓派生）的脚本：组合申报填表申报。
    SUPER_STRATEGY_COMBO_SCRIPTS = ("组合申报",)

    def __init__(self, root):
        self.root = root
        self.root.geometry("1000x780")
        self.root.minsize(850, 650)

        self.is_running = False
        self.current_category = "查询"
        self.current_script_name = None   # 当前选中的具体脚本（None 表示仅选中分类）

        # 自身引用（供任务中心等子模块访问主窗口能力）
        self.gui = self
        # 任务中心顺序执行模式开关
        self._task_mode = False
        self.task_center = None

        # 脚本列表 -> 任务队列 的拖拽状态
        self._drag_script = None
        self._drag_category = None   # 拖拽分类根节点时记录分类名
        self._drag_module = None     # 拖拽一级模块节点时记录模块名
        self._drag_active = False
        self._drag_occurred = False  # 本次按下是否实际发生了拖拽（用于决定是否跳过展开/折叠）
        self._drag_start_y = 0

        # 状态栏状态
        self._status_running = False
        self.task_start_time = 0.0

        # 加载用户配置
        self.user_config = load_user_config()

        # 当前客户端（多客户端支持）：优先读用户配置，否则取档案默认
        self.client_id = self.user_config.get("client") or get_default_client_id()
        if not get_client(self.client_id):
            self.client_id = get_default_client_id()
        self.client_var = tk.StringVar(value=get_client_name(self.client_id))

        # 参数变量（使用配置中的默认值）
        self.export_format = tk.StringVar(value=self.user_config.get("export_format", "txt"))
        self.auto_open = tk.BooleanVar(value=self.user_config.get("auto_open", False))
        self.log_level = tk.StringVar(value=self.user_config.get("log_level", "详细"))
        self.txt_path = tk.StringVar(value="")
        self.xls_path = tk.StringVar(value="")
        self.order_qty = tk.IntVar(value=1)
        self.countdown_sec = tk.IntVar(value=3)
        self.xlsx_file = tk.StringVar(value="")
        self.super_add_underlying = tk.BooleanVar(
            value=self.user_config.get("super_add_underlying", False)
        )
        configured_underlying = self.user_config.get(
            "super_strategy_underlying", DEFAULT_SUPER_STRATEGY_UNDERLYING
        )
        if configured_underlying not in SUPER_STRATEGY_UNDERLYINGS:
            configured_underlying = DEFAULT_SUPER_STRATEGY_UNDERLYING
        self.super_strategy_underlying = tk.StringVar(value=configured_underlying)

        # 超级策略 -> 组合申报 填表参数：市场 / 策略（均可多选打勾）/ 组合数量。
        # 合约一/合约二 不在此选择，由运行期按持仓派生。
        # 市场/策略用复选框（BooleanVar 字典），运行时按“市场 × 策略”逐个组合。
        self.super_combo_market_vars = {
            m: tk.BooleanVar(value=(m == SUPER_STRATEGY_COMBO_MARKETS[0]))
            for m in SUPER_STRATEGY_COMBO_MARKETS
        }
        self.super_combo_strategy_vars = {
            s: tk.BooleanVar(value=(s == SUPER_STRATEGY_COMBO_STRATEGIES[0]))
            for s in SUPER_STRATEGY_COMBO_STRATEGIES
        }
        self.super_combo_qty = tk.IntVar(value=1)

        # 期权下单_一键导出 参数
        self.export_target_position = tk.BooleanVar(value=True)  # 持仓
        self.export_target_order = tk.BooleanVar(value=True)     # 委托
        self.export_output_dir = tk.StringVar(value=get_output_dir(self.user_config, "下单"))

        # 交易系统设置 参数（输出路径可自定义）
        self.settings_output_dir = tk.StringVar(value=get_output_dir(self.user_config, "交易系统设置"))

        # 抓取自定义标准：勾选要抓取的面板（默认全选）
        self.capture_panels = {name: tk.BooleanVar(value=True) for name in CAPTURE_STANDARD_PANELS}

        # 日志目录：打包后放在exe同级目录
        if IS_FROZEN:
            self.log_dir = os.path.join(os.path.dirname(sys.executable), "logs")
        else:
            self.log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
        os.makedirs(self.log_dir, exist_ok=True)

        self._setup_logging()
        self._setup_runner()

        # 任务历史管理器（持久化到日志目录）
        self.history = HistoryManager(self.log_dir)
        self._current_record_id = None  # 当前正在运行的记录 id
        self._single_settings_batch = None
        self._single_stop_requested = False

        self._build_ui()
        self._update_title()
        self.logger.info("GUI自动化工具启动")

    def _setup_runner(self):
        """创建脚本执行引擎，并注入回调"""
        self.runner = ScriptRunner(
            is_frozen=IS_FROZEN,
            project_root=PROJECT_ROOT,
            log_level_getter=lambda: self.log_level.get(),
            on_log=self._log,
            on_debug=self.logger.debug,
            on_finish=self._on_run_finish,
            on_error=self._on_run_error,
        )

    def _setup_logging(self):
        """配置日志系统"""
        self.logger = logging.getLogger("AutomationGUI")
        self.logger.setLevel(logging.DEBUG)
        log_file = os.path.join(
            self.log_dir,
            f"gui_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        )
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        fmt = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(fmt)
        self.logger.addHandler(file_handler)

    def _build_ui(self):
        """构建界面"""
        # 菜单栏
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # 功能菜单（随客户端动态显示支持的分类）
        self.func_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="功能", menu=self.func_menu)
        self._rebuild_func_menu()

        # 工具菜单
        tool_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="工具", menu=tool_menu)
        tool_menu.add_command(label="清空日志", command=self._clear_log)
        tool_menu.add_command(label="打开日志目录", command=self._open_log_dir)
        tool_menu.add_separator()
        tool_menu.add_command(label="清空任务历史", command=self._clear_history)
        tool_menu.add_separator()
        tool_menu.add_command(label="退出", command=self.root.quit)

        # 设置菜单
        settings_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="设置", menu=settings_menu)
        settings_menu.add_radiobutton(
            label="详细日志",
            variable=self.log_level,
            value="详细",
            command=self._on_log_level_change
        )
        settings_menu.add_radiobutton(
            label="简洁日志",
            variable=self.log_level,
            value="简洁",
            command=self._on_log_level_change
        )
        settings_menu.add_separator()
        settings_menu.add_command(label="日志级别说明", command=self._show_log_level_help)

        # 主框架：用 grid 布局，row=0 可伸缩收缩，保证底部状态栏始终有空间
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky="nsew")

        # 顶部栏：客户端选择（多客户端支持）
        top_bar = ttk.Frame(main_frame)
        top_bar.pack(side=tk.TOP, fill=tk.X, pady=(0, 6))
        ttk.Label(top_bar, text="客户端:").pack(side=tk.LEFT, padx=(0, 4))
        self.client_combo = ttk.Combobox(
            top_bar, textvariable=self.client_var, state="readonly", width=18
        )
        self.client_combo["values"] = [get_client_name(cid) for cid in get_client_ids()]
        self.client_combo.pack(side=tk.LEFT)
        self.client_combo.bind("<<ComboboxSelected>>", self._on_client_change)

        # 标题
        ttk.Label(
            main_frame,
            text="钱龙期权交易自动化",
            font=("Microsoft YaHei UI", 16, "bold")
        ).pack(pady=(0, 10))

        # 左右分栏（PanedWindow 保证日志区域始终可见）
        self.paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        self.paned.pack(fill=tk.BOTH, expand=True)

        # 左侧面板：脚本列表 + 参数配置
        left_frame = ttk.Frame(self.paned)
        self.paned.add(left_frame, weight=1)

        # 操作按钮 - 固定在底部
        self.btn_frame = ttk.Frame(left_frame)
        self.btn_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(5, 0))

        self.execute_btn = ttk.Button(self.btn_frame, text="执行", command=self._execute_script, width=12)
        self.execute_btn.pack(side=tk.LEFT, padx=5)

        self.stop_btn = ttk.Button(self.btn_frame, text="停止", command=self._stop_script, state=tk.DISABLED, width=10)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        # 当前分类标签
        self.category_label = ttk.Label(
            left_frame,
            text="",
            font=("Microsoft YaHei UI", 12, "bold"),
            foreground="#0078d4"
        )
        self.category_label.pack(side=tk.TOP, anchor=tk.W, pady=(0, 5))

        # 中间内容区域 - 占据剩余空间
        middle_frame = ttk.Frame(left_frame)
        middle_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(0, 10))

        # 脚本列表（树形：分类 -> 脚本，随客户端动态显示/隐藏）
        list_frame = ttk.LabelFrame(middle_frame, text="功能 / 脚本", padding="8")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        self.script_tree = ttk.Treeview(
            list_frame,
            show="tree",
            selectmode=tk.BROWSE,
            height=12,
        )
        self.script_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.script_tree.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.script_tree.config(yscrollcommand=scrollbar.set)
        self.script_tree.bind('<Double-Button-1>', self._on_tree_double_click)
        self.script_tree.bind('<<TreeviewSelect>>', self._on_tree_select)
        # 从树拖入任务队列
        self.script_tree.bind('<ButtonPress-1>', self._on_list_drag_start)
        self.script_tree.bind('<B1-Motion>', self._on_list_drag_motion)
        self.script_tree.bind('<ButtonRelease-1>', self._on_list_drag_end)
        # 鼠标抬起时展开/折叠分类节点（避免按下即触发，与拖拽到队列冲突）
        self.script_tree.bind('<ButtonRelease-1>', self._on_tree_release, add='+')
        # iid -> {"script": 脚本配置, "category": 分类}
        self.tree_script_map = {}

        # 参数配置面板
        self.params_frame = ttk.LabelFrame(middle_frame, text="参数配置", padding="8")
        self.params_frame.pack(fill=tk.X, pady=(0, 10))

        # 数据预览面板（仅下单显示）
        self.preview_frame = ttk.LabelFrame(middle_frame, text="数据预览", padding="3")
        # 初始隐藏

        # 预览Treeview + 滚动条
        tree_wrap = ttk.Frame(self.preview_frame)
        tree_wrap.pack(fill=tk.X, anchor=tk.W)

        self.preview_tree = ttk.Treeview(tree_wrap, show="headings", height=5)
        v_scroll = ttk.Scrollbar(tree_wrap, orient=tk.VERTICAL, command=self.preview_tree.yview)
        h_scroll = ttk.Scrollbar(self.preview_frame, orient=tk.HORIZONTAL, command=self.preview_tree.xview)
        self.preview_tree.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)

        self.preview_tree.pack(side=tk.LEFT, fill=tk.BOTH)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)

        # 预览信息标签
        self.preview_info = ttk.Label(self.preview_frame, text="选择Excel文件后显示数据预览", foreground="gray")
        self.preview_info.pack(anchor=tk.W, pady=(3, 0))

        # 右侧：使用 Notebook 容纳「运行日志」与「任务历史」
        self.right_notebook = ttk.Notebook(self.paned)
        self.paned.add(self.right_notebook, weight=1)

        # —— 运行日志标签页 ——
        log_frame = ttk.Frame(self.right_notebook, padding="5")
        self.right_notebook.add(log_frame, text="运行日志")

        self.log_text = ColoredLogText(
            log_frame,
            font=("Consolas", 9),
            state=tk.DISABLED,
            wrap=tk.WORD,
            bg="#1e1e1e",
            fg="#d4d4d4",
            insertbackground="#d4d4d4",
            width=30
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # —— 任务历史标签页 ——
        history_frame = ttk.Frame(self.right_notebook, padding="5")
        self.right_notebook.add(history_frame, text="任务历史")
        self.history_panel = HistoryPanel(history_frame, self)

        # —— 任务中心标签页 ——
        task_frame = ttk.Frame(self.right_notebook, padding="5")
        self.right_notebook.add(task_frame, text="任务中心")
        self.task_center = TaskCenter(task_frame, self)
        # —— 定时任务标签页 ——
        sched_frame = ttk.Frame(self.right_notebook, padding="5")
        self.right_notebook.add(sched_frame, text="定时任务")
        self.scheduler_view = SchedulerPanel(sched_frame, self)
        # 创建定时任务调度器（后台线程）
        self.scheduler = TaskScheduler(self)
        self.scheduler_view.bind_scheduler(self.scheduler)

        # —— 结果比对待办标签页 ——
        compare_frame = ttk.Frame(self.right_notebook, padding="5")
        self.right_notebook.add(compare_frame, text="结果比对")
        self.compare_panel = ComparePanel(compare_frame, self)

        # —— 交易系统设置报告中心（位于"结果比对"右侧） ——
        report_frame = ttk.Frame(self.right_notebook, padding="5")
        self.right_notebook.add(report_frame, text="报告中心")
        self.report_center = SettingsReportPanel(report_frame, self)
        self._report_frame = report_frame

        # 状态栏：钉在窗口底部（始终可见，矮窗口也不会被裁切），
        # 但左右加 10px 内边距与主面板对齐，避免「分层」的割裂感
        status_outer = ttk.Frame(self.root, padding=(10, 0, 10, 0))
        status_outer.grid(row=1, column=0, sticky="ew")

        # 顶部细线分隔，替代原先加重的 SUNKEN 立体凹陷，与主面板风格保持一致
        ttk.Separator(status_outer, orient=tk.HORIZONTAL).pack(side=tk.TOP, fill=tk.X)

        status_frame = ttk.Frame(status_outer, padding=(2, 4))
        status_frame.pack(side=tk.TOP, fill=tk.X)

        self.status_label = ttk.Label(
            status_frame, text="就绪", anchor=tk.W
        )
        self.status_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.level_label = ttk.Label(
            status_frame, text=f"日志: {self.log_level.get()}", anchor=tk.CENTER,
            width=12
        )
        self.level_label.pack(side=tk.LEFT, padx=(8, 0))

        self.time_label = ttk.Label(
            status_frame, text="", anchor=tk.E, width=18
        )
        self.time_label.pack(side=tk.LEFT, padx=(8, 0))

        # 构建左侧「分类 -> 脚本」树，默认选中「下单」
        self._build_script_tree()
        self._select_category("下单")

        # 左右分栏比例：sash 位置占整体宽度的比例
        #   0.5 = 五五分 | 0.6 = 左6右4 | 0.8 = 左8右2（改这里即可调默认比例）
        self.pane_ratio = 0.5
        # 必须等窗口真正显示（<Map>）后再设置 sash，否则 winfo_width 为 1 -> 比例失效
        self.root.bind("<Map>", self._on_first_map)
        # 全局捕获鼠标释放，用于「从脚本列表拖入任务队列」的落点判定
        self.root.bind("<ButtonRelease-1>", self._on_global_drop)
        # 窗口缩放时按比例保持
        self.paned.bind("<Configure>", lambda e: self._apply_pane_ratio())

    def _on_first_map(self, event):
        """窗口首次显示后应用一次分栏比例，随后解绑"""
        self.root.unbind("<Map>")
        # 窗口完全显示后再布局分隔条
        self.root.after(20, self._apply_pane_ratio)
        self.root.after(100, self._apply_pane_ratio)

    def _apply_pane_ratio(self):
        """按固定比例设置左右分隔条位置（不受右侧标签页数量影响）"""
        w = self.paned.winfo_width()
        if w <= 1:
            return  # 窗口尚未绘制，宽度无效，跳过
        target = int(w * self.pane_ratio)
        if getattr(self, "_last_sash", None) == target:
            return  # 位置未变则跳过，避免与拖动/自身触发形成死循环
        self._last_sash = target
        self.paned.sashpos(0, target)

    def _build_script_tree(self):
        """重建左侧「一级模块 -> 分类 -> 脚本」树。"""
        self.script_tree.delete(*self.script_tree.get_children())
        self.tree_script_map.clear()
        scripts_config = get_scripts_config(self.client_id)
        for module, categories in MODULE_GROUPS.items():
            available = [c for c in categories if scripts_config.get(c)]
            if not available:
                continue
            module_iid = f"module::{module}"
            self.script_tree.insert(
                "", tk.END, iid=module_iid, text=module, open=(module == "行情交易")
            )
            for category in available:
                scripts = scripts_config[category]
                # 超级策略和交易系统设置本身就是一级模块，避免显示同名中间节点。
                if len(categories) == 1 and category == module:
                    parent_iid = module_iid
                else:
                    parent_iid = f"cat::{category}"
                    self.script_tree.insert(
                        module_iid, tk.END, iid=parent_iid, text=category, open=False
                    )
                for s in scripts:
                    # iid 必须唯一：查询类脚本共用驱动文件，故用 query_key 区分；
                    # 中泰等客户端脚本重定向后 path 可能重复，故再拼上脚本名。
                    sid_iid = (
                        f"script::{category}::{s.get('query_key') or s['path']}"
                        f"::{s['name']}"
                    )
                    self.script_tree.insert(parent_iid, tk.END, iid=sid_iid, text=s["name"])
                    self.tree_script_map[sid_iid] = {"script": s, "category": category}

    def _rebuild_func_menu(self):
        """重建「功能」菜单，按三个一级模块组织可用分类。"""
        self.func_menu.delete(0, tk.END)
        scripts_config = get_scripts_config(self.client_id)
        for module, categories in MODULE_GROUPS.items():
            available = [c for c in categories if scripts_config.get(c)]
            if not available:
                continue
            submenu = tk.Menu(self.func_menu, tearoff=0)
            for category in available:
                submenu.add_command(
                    label=category,
                    command=lambda c=category: self._select_category(c),
                )
            self.func_menu.add_cascade(label=module, menu=submenu)

    def _format_current_function(self, category, script_name=None):
        """拼出「当前功能」的完整面包屑：一级模块 / 分类 / 脚本（如选中具体脚本）"""
        module = get_module_for_category(category) or ""
        if module and module != category:
            parts = [module, category]
        else:
            parts = [category] if category else []
        if script_name:
            parts.append(script_name)
        return " / ".join(parts)

    def _select_category(self, category):
        """选中某个功能分类（来自树节点或功能菜单）"""
        self.current_category = category
        self.current_script_name = None
        self.category_label.config(text=f"当前功能: {self._format_current_function(category)}")
        # 仅展开所属一级模块；分类节点保持折叠，避免运行后默认展开（如「下单」）。
        module = get_module_for_category(category)
        module_iid = f"module::{module}"
        if module and self.script_tree.exists(module_iid):
            self.script_tree.item(module_iid, open=True)
            self.script_tree.update_idletasks()
            self.script_tree.after_idle(
                lambda i=module_iid: self.script_tree.see(i)
            )
            self.script_tree.after(
                80, lambda i=module_iid: self.script_tree.see(i)
            )
        # 清空脚本选择（避免沿用上一次选中的具体脚本）
        cur_sel = self.script_tree.selection()
        if cur_sel:
            self.script_tree.selection_remove(*cur_sel)
        self._rebuild_params()
        # 仅下单显示数据预览（其它分类隐藏并清空残留数据）
        self._update_preview_visibility(category == "下单")
        self._log(f"切换分类: {category}")
        self.logger.info(f"切换分类: {category}")
        # 空闲时同步状态栏（运行中不覆盖）
        if not self.is_running:
            self._set_status(f"就绪 - 当前功能: {self._format_current_function(category)}")

    def _on_tree_select(self, event=None):
        """树选择事件：脚本节点 -> 选中脚本；分类节点 -> 设为当前功能并展开"""
        sel = self.script_tree.selection()
        if not sel:
            return
        iid = sel[0]
        item = self.tree_script_map.get(iid)
        if item:
            # 选中具体脚本
            self.current_category = item["category"]
            self.current_script_name = item["script"]["name"]
            self.category_label.config(
                text=f"当前功能: {self._format_current_function(item['category'], item['script']['name'])}"
            )
            self._rebuild_params()
            self._update_paths_for_selected_script()
            self._update_params_for_selected_script()
        elif iid.startswith("cat::"):
            # 选中分类节点：设为当前功能（展开/折叠改到鼠标抬起时处理，见 _on_tree_release）
            category = iid.split("::", 1)[1]
            self.current_category = category
            self.current_script_name = None
            self.category_label.config(text=f"当前功能: {self._format_current_function(category)}")
            self._rebuild_params()
            # 仅下单显示数据预览（其它分类隐藏并清空残留数据）
            self._update_preview_visibility(category == "下单")
        else:
            # 一级模块节点只负责展开/折叠，不把模块名当成可执行分类；
            # 但仍更新「当前功能」标签，给出当前所在的模块层级反馈。
            module = iid.split("::", 1)[1]
            self.category_label.config(text=f"当前功能: {self._format_current_function(module)}")
            self._update_preview_visibility(False)
        # 确保选中项在视口中可见（_rebuild_params 会改变参数面板尺寸，
        # 进而压缩脚本树区域——list_frame 用 expand=True）。仅 after_idle 不够，
        # resize 通常在下一次事件循环才真正生效，需配合短延迟做兜底。
        self.script_tree.update_idletasks()
        self.script_tree.after_idle(lambda i=iid: self.script_tree.see(i))
        self.script_tree.after(80, lambda i=iid: self.script_tree.see(i))

    def _on_tree_double_click(self, event):
        """双击树：仅当双击具体脚本节点时才执行，双击分类/模块节点不执行，
        但需标记抑制紧接着的第二次释放翻转，使双击根节点等价于单击一次。"""
        iid = self.script_tree.identify_row(event.y)
        if not iid or iid not in self.tree_script_map:
            self._suppress_next_release = True
            return  # 分类/模块节点：不执行
        self._execute_script()

    # ====================== 任务历史标签页（已拆分为 gui/history_panel.py 的 HistoryPanel） ======================
    def _refresh_history(self):
        """刷新任务历史：转发给历史面板（任务中心/定时任务仍按此回调刷新）"""
        if hasattr(self, "history_panel"):
            self.history_panel.refresh()

    def _clear_history(self):
        """清空任务历史（菜单栏入口，转发给历史面板）"""
        if hasattr(self, "history_panel"):
            self.history_panel.clear_history()

    def _update_paths_for_selected_script(self):
        """根据当前选中的脚本更新路径显示"""
        script = self._get_selected_script()
        if not script:
            return
        output_dir = get_output_dir(self.user_config, self.current_category)
        filename = get_script_filename(script["name"])
        self.txt_path.set(os.path.join(output_dir, f"{filename}.txt"))
        self.xls_path.set(os.path.join(output_dir, f"{filename}.xls"))

    def _rebuild_params(self):
        """根据分类重建参数配置面板"""
        for w in self.params_frame.winfo_children():
            w.destroy()

        if self.current_category == "查询":
            self._build_query_params()
        elif self.current_category == "通知查询":
            self._build_query_params()
        elif self.current_category == "结算单":
            self._build_query_params()
        elif self.current_category == "下单":
            self._build_order_params()
        elif self.current_category == "组合申报":
            self._build_combo_params()
        elif self.current_category == "交易系统设置":
            if self._is_capture_script_selected():
                self._build_capture_params()
            else:
                self._build_settings_params()
        elif self.current_category in SUPER_STRATEGY_CATEGORIES:
            self._build_super_strategy_params()

    def _update_params_for_selected_script(self):
        """根据选中的脚本更新参数面板"""
        if self.current_category not in ("下单", "组合申报"):
            # 查询/通知查询/结算单/交易系统设置 等：不需要数据预览，隐藏并清空
            self._update_preview_visibility(False)
            return
        script = self._get_selected_script()
        if not script:
            return

        # 先清空参数配置面板
        for w in self.params_frame.winfo_children():
            w.destroy()

        if self.current_category == "下单":
            # 根据脚本决定是否显示数据预览
            order_script_name = get_script_filename(script["name"])
            if order_script_name in ("期权下单_一键导出", "全选撤单"):
                # 这两个脚本不需要数据预览
                self._update_preview_visibility(False)
            else:
                # 其他脚本需要数据预览
                self._update_preview_visibility(True)

            # 根据脚本更新参数配置面板
            if order_script_name == "全选撤单":
                # 全选撤单：显示提示信息
                ttk.Label(
                    self.params_frame,
                    text="该功能无需参数配置，点击执行即可。",
                    foreground="gray"
                ).pack(pady=20)
            elif order_script_name == "期权下单_一键导出":
                # 期权下单_一键导出：显示参数配置
                self._build_export_params()
            else:
                # 其他脚本：显示订单参数配置
                self._build_order_params()
        elif self.current_category == "组合申报":
            # 组合申报不需要数据预览，隐藏并清空残留数据
            self._update_preview_visibility(False)
            # 全自动脚本与查询脚本使用不同参数
            self._build_combo_params()

    def _build_query_params(self):
        """查询类参数"""
        ttk.Label(self.params_frame, text="导出格式:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(self.params_frame, text="TXT", variable=self.export_format, value="txt").grid(row=0, column=1, sticky=tk.W)
        ttk.Radiobutton(self.params_frame, text="XLS", variable=self.export_format, value="xls").grid(row=0, column=2, sticky=tk.W)

        ttk.Label(self.params_frame, text="自动打开文件:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Checkbutton(self.params_frame, text="导出后自动打开", variable=self.auto_open).grid(row=1, column=1, sticky=tk.W, columnspan=2)

        ttk.Label(self.params_frame, text="TXT输出路径:").grid(row=2, column=0, sticky=tk.W, pady=5)
        txt_entry = ttk.Entry(self.params_frame, textvariable=self.txt_path, width=35)
        txt_entry.grid(row=2, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=lambda: self._browse_path(self.txt_path, ".txt")).grid(row=2, column=2, padx=5)

        ttk.Label(self.params_frame, text="XLS输出路径:").grid(row=3, column=0, sticky=tk.W, pady=5)
        xls_entry = ttk.Entry(self.params_frame, textvariable=self.xls_path, width=35)
        xls_entry.grid(row=3, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=lambda: self._browse_path(self.xls_path, ".xls")).grid(row=3, column=2, padx=5)

        self.params_frame.columnconfigure(1, weight=1)

    def _build_order_params(self):
        """下单参数"""
        # 参数行
        param_row = ttk.Frame(self.params_frame)
        param_row.pack(fill=tk.X)

        ttk.Label(param_row, text="Excel文件:").grid(row=0, column=0, sticky=tk.W, pady=5)
        xlsx_entry = ttk.Entry(param_row, textvariable=self.xlsx_file, width=35)
        xlsx_entry.grid(row=0, column=1, sticky=tk.EW, pady=5)
        ttk.Button(param_row, text="选择文件", command=self._browse_xlsx).grid(row=0, column=2, padx=5)
        param_row.columnconfigure(1, weight=1)

        self.params_frame.columnconfigure(0, weight=1)

    def _build_combo_params(self):
        """组合申报参数：根据选中的脚本显示不同参数

        - 全自动脚本(组合申报_全自动/拆分申报_全自动): 仅需委托数量
        - 查询类脚本(组合策略持仓查询等): 导出格式/路径/自动打开
        """
        script = self._get_selected_script()
        if script and script["name"] in self.COMBO_AUTO_SCRIPTS:
            self._build_combo_auto_params()
        else:
            self._build_combo_query_params()

    def _build_combo_auto_params(self):
        """组合申报/拆分申报全自动：委托数量。"""
        ttk.Label(self.params_frame, text="委托数量:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Spinbox(self.params_frame, from_=1, to=999, textvariable=self.order_qty, width=10).grid(row=0, column=1, sticky=tk.W, pady=5)

        self.params_frame.columnconfigure(1, weight=1)

    def _build_combo_query_params(self):
        """组合策略查询类脚本：导出格式 / 路径 / 自动打开"""
        ttk.Label(self.params_frame, text="导出格式:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(self.params_frame, text="TXT", variable=self.export_format, value="txt").grid(row=0, column=1, sticky=tk.W)
        ttk.Radiobutton(self.params_frame, text="XLS", variable=self.export_format, value="xls").grid(row=0, column=2, sticky=tk.W)

        ttk.Label(self.params_frame, text="TXT输出路径:").grid(row=1, column=0, sticky=tk.W, pady=5)
        txt_entry = ttk.Entry(self.params_frame, textvariable=self.txt_path, width=35)
        txt_entry.grid(row=1, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=lambda: self._browse_path(self.txt_path, ".txt")).grid(row=1, column=2, padx=5)

        ttk.Label(self.params_frame, text="XLS输出路径:").grid(row=2, column=0, sticky=tk.W, pady=5)
        xls_entry = ttk.Entry(self.params_frame, textvariable=self.xls_path, width=35)
        xls_entry.grid(row=2, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=lambda: self._browse_path(self.xls_path, ".xls")).grid(row=2, column=2, padx=5)

        ttk.Label(self.params_frame, text="自动打开文件:").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Checkbutton(self.params_frame, text="导出后自动打开", variable=self.auto_open).grid(row=3, column=1, sticky=tk.W, columnspan=2)

        self.params_frame.columnconfigure(1, weight=1)

    def _build_settings_params(self):
        """交易系统设置参数 - 输出路径可自定义"""
        ttk.Label(self.params_frame, text="输出路径:").grid(row=0, column=0, sticky=tk.W, pady=5)
        path_entry = ttk.Entry(self.params_frame, textvariable=self.settings_output_dir, width=35)
        path_entry.grid(row=0, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=self._browse_settings_dir).grid(row=0, column=2, padx=5)

        ttk.Label(
            self.params_frame,
            text="测试报告与截图将保存到该目录下（reports / screenshots 子目录）",
            foreground="gray"
        ).grid(row=1, column=0, columnspan=3, sticky=tk.W, pady=5)

        self.params_frame.columnconfigure(1, weight=1)

    def _build_super_strategy_params(self):
        """超级策略参数：标的选择及是否在开仓前点击"加入标的"。

        组合申报需要选择市场/策略/组合数量（合约一/合约二按持仓派生，不在此选择）。
        """
        script = self._get_selected_script() if getattr(self, "script_tree", None) else None
        if script and script["name"] in self.SUPER_STRATEGY_COMBO_SCRIPTS:
            self._build_combination_declare_params()
            return
        ttk.Label(self.params_frame, text="超级策略标的:").grid(
            row=0, column=0, sticky=tk.W, pady=5
        )
        ttk.Combobox(
            self.params_frame,
            textvariable=self.super_strategy_underlying,
            values=SUPER_STRATEGY_UNDERLYINGS,
            state="readonly",
            width=28,
        ).grid(row=0, column=1, sticky=tk.W, pady=5)
        ttk.Checkbutton(
            self.params_frame,
            text="加入标的（可选）",
            variable=self.super_add_underlying,
        ).grid(row=1, column=0, columnspan=2, sticky=tk.W, pady=5)
        ttk.Label(
            self.params_frame,
            text=(
                "执行顺序：选择ETF标的 → 选择策略 → "
                "可选加入标的 → 一键开仓。"
            ),
            foreground="gray",
        ).grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=5)
        ttk.Label(
            self.params_frame,
            text="注意：选择加入标的请先手动登录。",
            foreground="#d32f2f",
            font=("Microsoft YaHei", 9, "bold"),
        ).grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=(0, 5))
        self.params_frame.columnconfigure(1, weight=1)

    def _build_combination_declare_params(self):
        """组合申报（超级策略分类）参数面板：市场 / 策略（复选）/ 组合数量。

        市场、策略均可用复选框打勾多选；运行时按“市场 × 策略”逐个打开组合申报对话框
        并组合。合约一 / 合约二 不在此选择——运行期会按对话框下拉候选项与所选策略
        从持仓派生配对（用户此前已用一键开仓建立对应组合）。
        """
        # ---------- 市场 ----------
        ttk.Label(self.params_frame, text="市场（可多选）:").grid(
            row=0, column=0, sticky=tk.NW, pady=5
        )
        market_frame = ttk.Frame(self.params_frame)
        market_frame.grid(row=0, column=1, sticky=tk.W, pady=5)
        for i, m in enumerate(SUPER_STRATEGY_COMBO_MARKETS):
            ttk.Checkbutton(
                market_frame, text=m, variable=self.super_combo_market_vars[m]
            ).grid(row=0, column=i, padx=(0, 10), sticky=tk.W)
        col = len(SUPER_STRATEGY_COMBO_MARKETS)
        ttk.Separator(market_frame, orient=tk.VERTICAL).grid(
            row=0, column=col, sticky=tk.NS, padx=(0, 6)
        )
        ttk.Button(
            market_frame, text="全选", width=4,
            command=lambda: self._toggle_combo_checkboxes("market", True),
        ).grid(row=0, column=col + 1, padx=(0, 2))
        ttk.Button(
            market_frame, text="全不选", width=5,
            command=lambda: self._toggle_combo_checkboxes("market", False),
        ).grid(row=0, column=col + 2)

        # ---------- 策略 ----------
        ttk.Label(self.params_frame, text="策略（可多选，逐个组合）:").grid(
            row=1, column=0, sticky=tk.NW, pady=5
        )
        strategy_frame = ttk.Frame(self.params_frame)
        strategy_frame.grid(row=1, column=1, sticky=tk.W, pady=5)
        for i, s in enumerate(SUPER_STRATEGY_COMBO_STRATEGIES):
            ttk.Checkbutton(
                strategy_frame, text=s, variable=self.super_combo_strategy_vars[s]
            ).grid(row=i, column=0, sticky=tk.W)
        # 全选/全不选按钮行
        action_row = len(SUPER_STRATEGY_COMBO_STRATEGIES)
        ttk.Separator(
            strategy_frame, orient=tk.HORIZONTAL
        ).grid(row=action_row, column=0, columnspan=2, sticky=tk.EW, pady=(4, 2))
        ttk.Button(
            strategy_frame, text="全选", width=4,
            command=lambda: self._toggle_combo_checkboxes("strategy", True),
        ).grid(row=action_row + 1, column=0, sticky=tk.W)
        ttk.Button(
            strategy_frame, text="全不选", width=5,
            command=lambda: self._toggle_combo_checkboxes("strategy", False),
        ).grid(row=action_row + 1, column=1, sticky=tk.W)

        ttk.Label(self.params_frame, text="组合数量:").grid(
            row=2, column=0, sticky=tk.W, pady=5
        )
        ttk.Entry(
            self.params_frame,
            textvariable=self.super_combo_qty,
            width=10,
        ).grid(row=2, column=1, sticky=tk.W, pady=5)

        ttk.Label(
            self.params_frame,
            text=(
                "勾选多个市场/策略时，会按“市场 × 策略”逐个打开组合申报并组合；\n"
                "合约一 / 合约二 由运行期按持仓派生，无需在此选择。"
            ),
            foreground="gray",
            justify=tk.LEFT,
        ).grid(row=3, column=0, columnspan=2, sticky=tk.W, pady=(8, 0))
        self.params_frame.columnconfigure(1, weight=1)

    def _toggle_combo_checkboxes(self, target: str, value: bool) -> None:
        """全选/全不选组合申报的复选框。"""
        vars_dict = (
            self.super_combo_market_vars
            if target == "market"
            else self.super_combo_strategy_vars
        )
        for var in vars_dict.values():
            var.set(value)

    def _is_capture_script_selected(self) -> bool:
        """当前选中的脚本是否为"抓取自定义标准"（按元数据标记判断，名字可随意改）"""
        script = self._get_selected_script()
        if not script:
            return False
        return bool(script.get("capture_standard"))

    def _capturable_panels(self):
        """当前客户端下可抓取的交易系统设置面板（按 clients.json 的 unsupported 动态过滤）。

        面板名与 SCRIPTS_CONFIG 的菜单名一致；与 get_scripts_config 同样的过滤规则：
        菜单名或 \\交易系统设置\\<面板名> 出现在客户端 unsupported 中即隐藏。
        例如钱龙客户端不支持"一键炒单设置"，参数配置里便不显示该项。
        """
        client = get_client(self.client_id) if self.client_id else None
        unsupported = set((client or {}).get("unsupported", []) or [])
        out = []
        for name in CAPTURE_STANDARD_PANELS:
            if name in unsupported:
                continue
            if f"\\交易系统设置\\{name}" in unsupported:
                continue
            out.append(name)
        return out

    def _build_capture_params(self):
        """抓取自定义标准 - 勾选要抓取的面板（按当前客户端动态过滤不支持的面板）"""
        panels = self._capturable_panels()

        # 标题行：左侧标签，右侧 全选 / 全不选
        head = ttk.Frame(self.params_frame)
        head.grid(row=0, column=0, columnspan=4, sticky=tk.EW, pady=(2, 8))
        head.columnconfigure(0, weight=1)
        ttk.Label(head, text="选择要抓取的面板:").grid(row=0, column=0, sticky=tk.W)
        btn_row = ttk.Frame(head)
        btn_row.grid(row=0, column=1, sticky=tk.E)
        ttk.Button(
            btn_row, text="全选",
            command=lambda: self._set_capture_all(True),
        ).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(
            btn_row, text="全不选",
            command=lambda: self._set_capture_all(False),
        ).pack(side=tk.LEFT)

        if not panels:
            ttk.Label(
                self.params_frame,
                text="当前客户端不支持任何可抓取的面板。",
                foreground="gray",
            ).grid(row=1, column=0, sticky=tk.W, pady=6, columnspan=4)
            return

        # 勾选框（每行 2 个）
        cols = 2
        for idx, name in enumerate(panels):
            var = self.capture_panels.get(name)
            if var is None:
                var = tk.BooleanVar(value=True)
                self.capture_panels[name] = var
            r, c = divmod(idx, cols)
            ttk.Checkbutton(self.params_frame, text=name, variable=var).grid(
                row=1 + r, column=c, sticky=tk.W, padx=(0, 20), pady=3
            )

        # 操作按钮：打开标准文件夹 / 恢复默认
        checkbox_rows = (len(panels) + cols - 1) // cols
        act_row = ttk.Frame(self.params_frame)
        act_row.grid(row=1 + checkbox_rows, column=0, columnspan=4, sticky=tk.W, pady=(12, 0))
        ttk.Button(
            act_row, text="打开标准文件夹",
            command=self._open_standard_folder,
        ).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(
            act_row, text="恢复默认(删除选中面板JSON)",
            command=self._restore_default_standards,
        ).pack(side=tk.LEFT)

    def _set_capture_all(self, value: bool):
        """全选 / 全不选（仅作用于当前客户端可见的面板）"""
        for name in self._capturable_panels():
            var = self.capture_panels.get(name)
            if var is not None:
                var.set(value)

    def _open_standard_folder(self):
        """打开当前客户端的标准目录，便于查看/手动删除 JSON。

        客户端专属目录不存在时回退到标准根目录；根目录也不存在时提示尚未抓取。
        """
        client_dir = STANDARD_ROOT / (self.client_id or "default")
        try:
            open_path(client_dir)
            return
        except FileNotFoundError:
            pass
        try:
            open_path(STANDARD_ROOT)
        except FileNotFoundError:
            messagebox.showinfo("提示", "标准目录尚不存在（尚未抓取过任何自定义标准）。")

    def _restore_default_standards(self):
        """删除勾选面板在当前客户端的自定义标准 JSON，恢复到默认/兜底标准。

        删除前自动备份为 .json.bak；若该面板本就无自定义 JSON，则提示已是默认状态。
        删除后 load_standard 会回退到 default 兜底目录或脚本内嵌默认值。
        """
        selected = [
            n for n in self._capturable_panels()
            if self.capture_panels.get(n) and self.capture_panels[n].get()
        ]
        if not selected:
            messagebox.showwarning("提示", "请先勾选要恢复默认的面板")
            return
        client_name = get_client_name(self.client_id)
        detail = "\n".join(f"  · {n}" for n in selected)
        ans = messagebox.askyesno(
            "确认恢复默认",
            f"将删除以下 {len(selected)} 个面板在当前客户端（{client_name}）的\n"
            f"自定义标准 JSON，恢复到默认/兜底标准：\n\n{detail}\n\n"
            f"（删除前会自动备份为 .json.bak，可手动还原）",
        )
        if not ans:
            return
        deleted = []
        for name in selected:
            path = resolve_standard_path(name, self.client_id or "default")
            if path.is_file():
                try:
                    shutil.copy2(path, path.with_suffix(".json.bak"))
                    path.unlink()
                    deleted.append(name)
                except Exception as exc:  # noqa: BLE001
                    messagebox.showerror("错误", f"删除 {name} 标准失败: {exc}")
        if deleted:
            messagebox.showinfo("完成", f"已恢复默认标准：{', '.join(deleted)}")
        else:
            messagebox.showinfo("完成", "所选面板当前没有自定义标准，已是默认状态。")

    def _build_export_params(self):
        """期权下单_一键导出 参数配置"""
        # 清空旧内容
        for widget in self.params_frame.winfo_children():
            widget.destroy()

        # 导出目标选择
        ttk.Label(self.params_frame, text="导出目标:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Checkbutton(self.params_frame, text="持仓", variable=self.export_target_position).grid(row=0, column=1, sticky=tk.W, padx=5)
        ttk.Checkbutton(self.params_frame, text="委托", variable=self.export_target_order).grid(row=0, column=2, sticky=tk.W, padx=5)

        # 输出目录配置
        ttk.Label(self.params_frame, text="输出目录:").grid(row=1, column=0, sticky=tk.W, pady=5)
        path_entry = ttk.Entry(self.params_frame, textvariable=self.export_output_dir, width=35)
        path_entry.grid(row=1, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=self._browse_export_dir).grid(row=1, column=2, padx=5)

        # 显示文件名格式说明
        ttk.Label(self.params_frame, text="文件名格式: 期权下单(新)-持仓-20260629.xls", foreground="gray").grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=5)

        self.params_frame.columnconfigure(1, weight=1)

        # 确保按钮框架可见
        self.btn_frame.lift()

    def _browse_path(self, var, ext):
        """浏览选择文件路径"""
        initial_dir = get_output_dir(self.user_config, self.current_category)
        # 默认文件名：当前输入框已有值，或根据选中脚本生成
        current = var.get()
        if current and os.path.basename(current):
            initial_file = os.path.basename(current)
        else:
            script = self._get_selected_script()
            if script:
                initial_file = get_script_filename(script["name"]) + ext
            else:
                initial_file = "output" + ext

        path = filedialog.asksaveasfilename(
            initialdir=initial_dir,
            initialfile=initial_file,
            defaultextension=ext,
            filetypes=[("All files", "*.*"), (f"{ext.upper()} files", f"*{ext}")]
        )
        if path:
            var.set(path)
            new_dir = os.path.dirname(path)
            if new_dir != get_output_dir(self.user_config, self.current_category):
                set_output_dir(self.user_config, self.current_category, new_dir)
                save_user_config(self.user_config)
                self._log(f"[配置] 已更新{self.current_category}输出目录: {new_dir}")

    def _browse_xlsx(self):
        """选择Excel文件"""
        path = filedialog.askopenfilename(
            title="选择Excel配置文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self.xlsx_file.set(path)
            self._log(f"已选择Excel文件: {path}")
            self._preview_excel(path)

    def _browse_export_dir(self):
        """选择导出目录"""
        path = filedialog.askdirectory(
            title="选择导出目录",
            initialdir=self.export_output_dir.get()
        )
        if path:
            self.export_output_dir.set(path)
            self._log(f"已设置导出目录: {path}")

    def _browse_settings_dir(self):
        """选择交易系统设置输出目录"""
        path = filedialog.askdirectory(
            title="选择输出目录",
            initialdir=self.settings_output_dir.get()
        )
        if path:
            self.settings_output_dir.set(path)
            set_output_dir(self.user_config, "交易系统设置", path)
            save_user_config(self.user_config)
            self._log(f"[配置] 已更新交易系统设置输出目录: {path}")
            if hasattr(self, "report_center"):
                self.report_center.refresh_batches()

    def _preview_excel(self, filepath: str):
        """读取Excel并显示到预览表格"""
        if not HAS_OPENPYXL:
            self.preview_info.config(text="未安装 openpyxl，无法预览。请运行: pip install openpyxl", foreground="red")
            return

        # 清空旧数据
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = ()

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active

            # 使用 iter_rows 安全读取，避免直接访问 ws[1] 导致索引错误
            rows_iter = ws.iter_rows(values_only=True)

            # 读取第一行作为表头
            try:
                first_row = next(rows_iter)
            except StopIteration:
                self.preview_info.config(text="Excel文件无任何数据", foreground="red")
                wb.close()
                return

            if not first_row or all(v is None for v in first_row):
                self.preview_info.config(text="Excel文件无表头或为空", foreground="red")
                wb.close()
                return

            headers = [str(v) if v is not None else "" for v in first_row]

            if all(h == "" for h in headers):
                self.preview_info.config(text="Excel文件无表头或为空", foreground="red")
                wb.close()
                return

            # 设置列（紧凑宽度，不拉伸）
            self.preview_tree["columns"] = headers
            for h in headers:
                self.preview_tree.heading(h, text=h)
                self.preview_tree.column(h, width=64, minwidth=30, stretch=False)

            # 读取数据行（最多显示50行）
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(v is None for v in row):
                    continue
                vals = [str(v) if v is not None else "" for v in row]
                # 补齐列数
                while len(vals) < len(headers):
                    vals.append("")
                self.preview_tree.insert("", tk.END, values=vals[:len(headers)])
                row_count += 1
                if row_count >= 50:
                    break

            total_rows = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and any(v is not None for v in r))
            show_note = f"（共{total_rows}行，仅显示前50行）" if total_rows > 50 else f"（共{total_rows}行）"
            self.preview_info.config(text=f"字段: {len(headers)} 个 | 数据: {show_note}", foreground="green")

            wb.close()

        except Exception as e:
            self.preview_info.config(text=f"读取失败: {e}", foreground="red")
            self._log(f"[预览] Excel读取失败: {e}")

    def _update_preview_visibility(self, show):
        """显示/隐藏数据预览面板。

        隐藏时同时清空已加载的预览数据与 Excel 文件选择，
        避免切换到其它脚本（如查询）时残留旧数据预览。
        """
        if show:
            self.preview_frame.pack(fill=tk.X, pady=(0, 10))
            return
        # 隐藏面板并清空残留数据，防止切回其它脚本后旧预览仍可见
        self.preview_frame.pack_forget()
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = ()
        self.preview_info.config(text="选择Excel文件后显示数据预览", foreground="gray")
        self.xlsx_file.set("")

    def _get_selected_script(self):
        """获取选中的脚本（仅当树中选中具体脚本节点时返回）"""
        sel = self.script_tree.selection()
        if not sel:
            return None
        item = self.tree_script_map.get(sel[0])
        return item["script"] if item else None

    # ====================== 脚本列表 -> 任务队列 拖拽 ======================
    def _on_list_drag_start(self, event):
        """从树中开始拖拽（记录待拖出的脚本、分类或一级模块）"""
        self._drag_script = None
        self._drag_category = None
        self._drag_module = None
        self._drag_active = False
        self._drag_occurred = False
        self._drag_iid = None
        self._suppress_next_release = False
        if self.task_center is None or self.task_center.is_running:
            return
        iid = self.script_tree.identify_row(event.y)
        if not iid:
            return
        self._drag_iid = iid
        # 一级模块（iid 形如 module::行情交易）：落点时加入模块下全部可用分类。
        if iid.startswith("module::"):
            module = iid.split("::", 1)[1]
            if module in MODULE_GROUPS:
                self._drag_module = module
                self._drag_start_y = event.y
            return
        # 拖拽分类根节点（iid 形如 cat::查询）：记录分类名，落点时加入其下全部脚本
        if iid.startswith("cat::"):
            self._drag_category = iid.split("::", 1)[1]
            self._drag_start_y = event.y
            return
        if iid not in self.tree_script_map:
            return
        item = self.tree_script_map[iid]
        self._drag_script = dict(item["script"])
        self._drag_script["category"] = item["category"]
        self._drag_start_y = event.y

    def _on_list_drag_motion(self, event):
        """拖动过程中：超过阈值视为拖拽，并在悬停于队列时显示落点"""
        if (self._drag_script is None and self._drag_category is None
                and self._drag_module is None):
            return
        if abs(event.y - self._drag_start_y) < 6:
            return
        self._drag_active = True
        self._drag_occurred = True
        tc = self.task_center
        if tc is None:
            return
        tree = tc.tree
        rx, ry = tree.winfo_rootx(), tree.winfo_rooty()
        if rx <= event.x_root <= rx + tree.winfo_width() and ry <= event.y_root <= ry + tree.winfo_height():
            tc._update_drop_indicator(event.y_root - ry)
        else:
            tc._hide_drop_indicator()

    def _on_list_drag_end(self, event):
        """列表框内释放：若不在队列上方则取消拖拽状态"""
        tc = self.task_center
        over_tree = False
        if tc is not None:
            tree = tc.tree
            rx, ry = tree.winfo_rootx(), tree.winfo_rooty()
            over_tree = (rx <= event.x_root <= rx + tree.winfo_width()
                         and ry <= event.y_root <= ry + tree.winfo_height())
        if not over_tree:
            self._drag_script = None
            self._drag_category = None
            self._drag_module = None
            self._drag_active = False
            if tc is not None:
                tc._hide_drop_indicator()
        # 在队列上方释放时保留状态，交由 _on_global_drop 处理落点与清理

    def _on_global_drop(self, event):
        """全局捕获释放：处理从脚本列表拖入队列的落点。"""
        script = self._drag_script
        category = self._drag_category
        module = self._drag_module
        active = self._drag_active
        self._drag_script = None
        self._drag_category = None
        self._drag_module = None
        self._drag_active = False
        tc = self.task_center
        if tc is not None:
            tc._hide_drop_indicator()
        if not active or tc is None:
            return
        if tc.is_running:
            return
        tree = tc.tree
        rx, ry = tree.winfo_rootx(), tree.winfo_rooty()
        if rx <= event.x_root <= rx + tree.winfo_width() and ry <= event.y_root <= ry + tree.winfo_height():
            if module:
                tc.add_module_from_drop(module, event.y_root - ry)
            elif category:
                tc.add_category_from_drop(category, event.y_root - ry)
            elif script is not None:
                tc.add_script_from_drop(script, event.y_root - ry)

    def _on_tree_release(self, event):
        """鼠标抬起时展开/折叠分类节点

        仅在「本次按下没有发生拖拽」时切换展开状态，
        避免从脚本列表拖拽到任务队列时误触发展开/折叠。
        """
        if getattr(self, "_drag_occurred", False):
            self._drag_occurred = False
            return
        # 用按下时记录的节点（而非释放时重新识别），避免切换分类触发预览面板
        # 收起导致布局回流、释放坐标错位而无法展开节点。
        iid = getattr(self, "_drag_iid", None)
        if not iid or not iid.startswith(("cat::", "module::")):
            return
        # 仅当抬起位置就是当前选中的分类节点时才切换
        sel = self.script_tree.selection()
        if not sel or sel[0] != iid:
            return
        # 双击根节点时，第二次按下会触发 <Double-Button-1> 并标记抑制，
        # 此处跳过本次翻转，使双击等价于单击一次（避免两次翻转互相抵消）。
        if getattr(self, "_suppress_next_release", False):
            self._suppress_next_release = False
            return
        self.script_tree.item(iid, open=not self.script_tree.item(iid, "open"))
        # 展开/折叠后确保当前节点在视口中可见（idle 队列处理避免布局未稳定）
        self.script_tree.update_idletasks()
        self.script_tree.after_idle(lambda i=iid: self.script_tree.see(i))
        self.script_tree.after(80, lambda i=iid: self.script_tree.see(i))

    def _execute_script(self):
        """执行脚本"""
        if self.is_running:
            messagebox.showwarning("提示", "有脚本正在运行中，请先停止")
            return
        if self._task_mode:
            messagebox.showwarning("提示", "任务中心正在顺序执行中，请先停止")
            return

        script = self._get_selected_script()
        if not script:
            messagebox.showwarning("提示", "请先选择一个脚本")
            return

        if not os.path.exists(script["path"]):
            messagebox.showerror("错误", f"脚本文件不存在:\n{script['path']}")
            return

        # 下单需要检查Excel文件（全选撤单和期权下单_一键导出除外）
        order_script_name = get_script_filename(script["name"])
        if self.current_category == "下单" and order_script_name not in ("期权下单_一键导出", "全选撤单") and not self.xlsx_file.get():
            messagebox.showwarning("提示", "请先选择Excel配置文件")
            return

        # 抓取自定义标准需要至少勾选一个面板
        if self.current_category == "交易系统设置" and self._is_capture_script_selected():
            selected_panels = [n for n in self._capturable_panels()
                               if self.capture_panels.get(n) and self.capture_panels[n].get()]
            if not selected_panels:
                messagebox.showwarning("提示", "请至少勾选一个要抓取的面板")
                return

        # 期权下单_一键导出需要检查导出目标
        export_targets = []
        if order_script_name == "期权下单_一键导出":
            if self.export_target_position.get():
                export_targets.append("持仓")
            if self.export_target_order.get():
                export_targets.append("委托")
            if not export_targets:
                messagebox.showwarning("提示", "请至少选择一个导出目标（持仓或委托）")
                return

        # 保存当前配置
        self.user_config["export_format"] = self.export_format.get()
        self.user_config["auto_open"] = self.auto_open.get()
        self.user_config["super_add_underlying"] = self.super_add_underlying.get()
        self.user_config["super_strategy_underlying"] = (
            self.super_strategy_underlying.get()
        )
        if self.current_category == "交易系统设置":
            set_output_dir(self.user_config, "交易系统设置", self.settings_output_dir.get())
        save_user_config(self.user_config)

        # 确保路径已设置
        if self.current_category in ("查询", "通知查询", "结算单") and not self.txt_path.get():
            self._update_paths_for_selected_script()

        # 收集运行时参数，构造任务
        params = self.collect_params(export_targets)
        if self.current_category == "交易系统设置":
            if self._is_capture_script_selected():
                # 抓取自定义标准是只读采集，不产生比对差异，
                # 不创建报告中心批次，运行结果也不在报告中心显示。
                self._single_settings_batch = None
            else:
                batch_item = {
                    "category": "交易系统设置",
                    "script_name": script["name"],
                    "script_path": script["path"],
                    "params": params,
                    "status": TaskCenter.ST_PENDING,
                }
                try:
                    self._single_settings_batch = self.begin_settings_batch(
                        "单独运行", [batch_item]
                    )
                except OSError as exc:
                    messagebox.showerror("报告中心", f"无法创建报告批次:\n{exc}")
                    return
                params.update(batch_item.get("_settings_runtime", {}))
                self._single_stop_requested = False
        task = Task(script, self.current_category, params)

        self.is_running = True
        self.execute_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        self._set_status(f"正在运行: {script['name']}", running=True)

        self._log(f"\n{'='*60}")
        self._log(f"开始执行: {script['name']}")
        self._log(f"脚本路径: {script['path']}")

        # 打印参数
        if self.current_category == "查询":
            self._log(f"导出格式: {self.export_format.get().upper()}")
            self._log(f"自动打开: {'是' if self.auto_open.get() else '否'}")
            self._log(f"TXT路径: {self.txt_path.get()}")
            self._log(f"XLS路径: {self.xls_path.get()}")
        elif self.current_category == "通知查询":
            self._log(f"导出格式: {self.export_format.get().upper()}")
            self._log(f"自动打开: {'是' if self.auto_open.get() else '否'}")
            self._log(f"TXT路径: {self.txt_path.get()}")
            self._log(f"XLS路径: {self.xls_path.get()}")
        elif self.current_category == "结算单":
            self._log(f"导出格式: {self.export_format.get().upper()}")
            self._log(f"自动打开: {'是' if self.auto_open.get() else '否'}")
            self._log(f"TXT路径: {self.txt_path.get()}")
            self._log(f"XLS路径: {self.xls_path.get()}")
        elif self.current_category == "下单":
            if order_script_name == "期权下单_一键导出":
                self._log(f"导出目标: {', '.join(export_targets)}")
                self._log(f"输出目录: {self.export_output_dir.get()}")
                self._log(f"文件名格式: 期权下单(新)-持仓-20260629.xls")
            else:
                self._log(f"Excel文件: {self.xlsx_file.get()}")
        elif self.current_category == "组合申报":
            if script["name"] in self.COMBO_AUTO_SCRIPTS:
                self._log(f"委托数量: {self.order_qty.get()}")
                self._log("运行模式: 正式申报（全量遍历）")
            else:
                self._log(f"导出格式: {self.export_format.get().upper()}")
                self._log(f"自动打开: {'是' if self.auto_open.get() else '否'}")
                self._log(f"TXT路径: {self.txt_path.get()}")
                self._log(f"XLS路径: {self.xls_path.get()}")
        elif self.current_category == "交易系统设置":
            if self._is_capture_script_selected():
                self._log(f"抓取面板: {', '.join(n for n in self._capturable_panels() if self.capture_panels.get(n) and self.capture_panels[n].get())}")
            else:
                self._log(f"输出路径: {self.settings_output_dir.get()}")
        elif self.current_category in SUPER_STRATEGY_CATEGORIES:
            if script["name"] in self.SUPER_STRATEGY_COMBO_SCRIPTS:
                markets = [m for m, v in self.super_combo_market_vars.items() if v.get()]
                strategies = [s for s, v in self.super_combo_strategy_vars.items() if v.get()]
                self._log(f"市场: {', '.join(markets) or '(未选)'}")
                self._log(f"策略: {', '.join(strategies) or '(未选)'}")
                self._log(f"组合数量: {self.super_combo_qty.get()}")
                self._log(f"将依次组合 {len(markets) * len(strategies)} 个组合")
                self._log("合约一/合约二: 运行期按持仓派生")
            else:
                self._log(f"超级策略标的: {self.super_strategy_underlying.get()}")
                self._log(
                    f"加入标的: {'是' if self.super_add_underlying.get() else '否'}"
                )
                self._log("下单动作: 一键开仓")

        self._log(f"{'='*60}")
        self.logger.info(f"开始执行: {script['name']}")

        # 交给执行引擎在后台线程运行
        self._current_record_id = self.history.add_record(script["name"], self.current_category)
        self.history_panel.reset_page()  # 新任务置顶，回到第 1 页并刷新
        self.runner.run(task)

    def collect_params(self, export_targets=None):
        """收集当前界面参数，返回 dict（供执行/任务中心使用）"""
        return {
            "export_format": self.export_format.get(),
            "auto_open": self.auto_open.get(),
            "txt_path": self.txt_path.get(),
            "xls_path": self.xls_path.get(),
            "order_qty": self.order_qty.get(),
            "countdown_sec": self.countdown_sec.get(),
            "xlsx_file": self.xlsx_file.get(),
            "export_targets": export_targets or [],
            "export_output_dir": self.export_output_dir.get(),
            "settings_output_dir": self.settings_output_dir.get(),
            "super_add_underlying": self.super_add_underlying.get(),
            "super_strategy_underlying": self.super_strategy_underlying.get(),
            "super_combo_markets": [
                m for m, v in self.super_combo_market_vars.items() if v.get()
            ],
            "super_combo_strategies": [
                s for s, v in self.super_combo_strategy_vars.items() if v.get()
            ],
            "super_combo_qty": self.super_combo_qty.get(),
            "client_id": self.client_id,
            "capture_panels": [n for n in self._capturable_panels()
                               if self.capture_panels.get(n) and self.capture_panels[n].get()],
        }

    def make_task_params(self, script, category):
        """为指定脚本生成任务参数快照（路径按脚本自身名称分别生成默认路径）

        用于批量加入整分类时，让每个脚本携带各自的默认输出路径，
        而非共享拖拽瞬间的同一份界面参数。
        """
        params = self.collect_params()
        # 查询类 / 结算单 / 组合申报中的查询脚本：输出路径按脚本名称分别生成
        if category in ("查询", "通知查询", "结算单"):
            output_dir = get_output_dir(self.user_config, category)
            filename = get_script_filename(script["name"])
            params["txt_path"] = os.path.join(output_dir, f"{filename}.txt")
            params["xls_path"] = os.path.join(output_dir, f"{filename}.xls")
        elif category == "组合申报" and script["name"] not in self.COMBO_AUTO_SCRIPTS:
            output_dir = get_output_dir(self.user_config, category)
            filename = get_script_filename(script["name"])
            params["txt_path"] = os.path.join(output_dir, f"{filename}.txt")
            params["xls_path"] = os.path.join(output_dir, f"{filename}.xls")
        return params

    # ====================== 客户端切换（多客户端支持） ======================
    def _update_title(self):
        """根据当前客户端更新窗口标题"""
        self.root.title(f"{get_client_name(self.client_id)} - GUI自动化工具")

    def _on_client_change(self, event=None):
        """下拉切换客户端：持久化配置并刷新当前分类脚本列表"""
        name = self.client_var.get()
        cid = next((c for c in get_client_ids() if get_client_name(c) == name), None)
        if not cid or cid == self.client_id:
            return
        self.client_id = cid
        self.user_config["client"] = cid
        save_user_config(self.user_config)
        self._update_title()
        self._log(f"[配置] 已切换客户端: {name}")
        self.logger.info(f"切换客户端: {cid}")
        # 重建树与功能菜单（应用 unsupported 过滤），并复位到有效分类
        self._build_script_tree()
        self._rebuild_func_menu()
        supported = [c for c in CATEGORIES if get_scripts_config(self.client_id).get(c)]
        if self.current_category not in supported:
            self.current_category = supported[0] if supported else ""
        if self.current_category:
            self._select_category(self.current_category)
        if hasattr(self, "report_center"):
            self.report_center.refresh_batches()
        # 任务中心编队按客户端过滤：切换客户端后刷新编队下拉（复位到「自定义队列」）
        if self.task_center is not None:
            self.task_center._refresh_group_combo()

    def show_report_center(self, auto_clear=True):
        """切换到交易系统设置报告中心标签页。

        面板仅在软件打开时清空一次（见 SettingsReportPanel.__init__ 的
        refresh_batches），切换标签页不再清空，以保留用户正在查看的批次结果。
        auto_clear 仅保留作兼容参数，已无实际清空作用。
        """
        if not hasattr(self, "_report_frame"):
            return
        self.right_notebook.select(self._report_frame)

    # ====================== 交易系统设置：统一批次报告 ======================
    def begin_settings_batch(self, source, task_items):
        """为一次运行创建批次，并给其中的设置任务附加运行上下文。"""
        settings_tasks = [
            item for item in task_items
            if item.get("category") == "交易系统设置"
        ]
        if not settings_tasks:
            return None

        output_dir = self.settings_output_dir.get().strip()
        if not output_dir:
            raise OSError("交易系统设置输出目录为空")
        run_id = create_run_id()
        batch_dir = os.path.abspath(os.path.join(output_dir, "批次", run_id))
        os.makedirs(batch_dir, exist_ok=False)
        context = {
            "run_id": run_id,
            "batch_dir": batch_dir,
            "output_dir": output_dir,
            "client_id": self.client_id,
            "source": source,
            "summary": None,
        }
        runtime = {
            "settings_output_dir": output_dir,
            "settings_run_id": run_id,
            "settings_run_dir": batch_dir,
            "client_id": self.client_id,
        }
        for item in settings_tasks:
            item["_settings_runtime"] = dict(runtime)

        self._log(
            f"[报告中心] 已创建{source}批次: {run_id} | "
            f"设置模块 {len(settings_tasks)} 个"
        )
        return context

    def update_settings_batch(self, context, task_records, final=False, stopped=False):
        """任务结束时生成一次总报告，并把完整批次信息直接交给报告中心。"""
        if not context:
            return None
        if not final:
            return None
        records = [
            {**item, "params": dict(item.get("params", {}))}
            for item in task_records
            if item.get("category") == "交易系统设置"
        ]
        batch_status = (
            BATCH_STOPPED if stopped
            else BATCH_COMPLETED
        )
        try:
            summary = generate_batch_reports(
                run_id=context["run_id"],
                batch_dir=context["batch_dir"],
                client_id=context["client_id"],
                task_records=records,
                stopped=stopped,
                source=context["source"],
                batch_status=batch_status,
            )
        except Exception as exc:
            self._log(f"[报告中心] 批次报告更新失败: {exc}")
            self.logger.exception("交易系统设置批次报告更新失败")
            return None

        context["summary"] = summary
        if hasattr(self, "report_center"):
            self.report_center.on_batch_summary(summary, final=final)
        return summary

    def resume_settings_batch(self, context, task_records):
        """停止后继续执行时，沿用原批次并恢复为运行中。"""
        runtime = {
            "settings_output_dir": context["output_dir"],
            "settings_run_id": context["run_id"],
            "settings_run_dir": context["batch_dir"],
            "client_id": context["client_id"],
        }
        for item in task_records:
            if item.get("category") == "交易系统设置":
                item["_settings_runtime"] = dict(runtime)
        return None

    # ====================== 任务中心：顺序执行驱动 ======================
    def run_task_center(self, task_center):
        """由任务中心调用：进入顺序执行模式并启动首个任务"""
        self._task_mode = True
        self.task_center = task_center
        # 禁用主窗口执行/停止按钮，避免与任务中心冲突
        self.execute_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.DISABLED)
        task_center.run_next()

    def stop_task_center(self):
        """由任务中心调用：停止当前正在执行的子进程"""
        if self.runner.is_running:
            self._log("\n[任务中心] 正在终止当前子进程...")
            self.logger.info("任务中心 - 用户手动停止")
            self.runner.stop()
        # 子进程结束后会触发回调，task_center 据此进入停止收尾流程

    def execute_task_item(self, item, record_id, next_category=""):
        """由任务中心调用：执行队列中的单个任务项（带参数快照）

        next_category: 下一个任务的分类，传给脚本用于决定交易系统设置窗口是否保留。
        """
        script = {"name": item["script_name"], "path": item["script_path"],
                  "query_key": item.get("query_key", "")}
        params = dict(item["params"])
        params.update(item.get("_settings_runtime", {}))
        task = Task(script, item["category"], params, next_category=next_category)
        self.runner.run(task)

    def _reset_running_state_if_idle(self):
        """任务中心收尾时复位主窗口运行状态（若非普通执行占用）"""
        self._task_mode = False
        self._reset_running_state()

    def _stop_script(self):
        """停止脚本"""
        if self.runner.is_running:
            if not self._task_mode:
                self._single_stop_requested = True
            self._log("\n[停止] 用户手动停止...")
            self.logger.info("用户手动停止")
            self._set_status("已停止（用户手动）")

            if self._current_record_id is not None:
                self.history.update_record(self._current_record_id, STATUS_STOPPED)
                self._current_record_id = None
                self._refresh_history()

            self.runner.stop()

    def _clear_log(self):
        """清空日志"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

    def _open_log_dir(self):
        """打开日志目录"""
        try:
            open_path(self.log_dir)
        except OSError as exc:
            messagebox.showerror("打开失败", str(exc))

    # ====================== 执行结果回调（运行在 runner 线程，统一切回主线程更新 UI） ======================
    def _on_run_finish(self, return_code, elapsed, task):
        login_message = getattr(task, "login_required_message", "")
        trading_time_message = getattr(task, "trading_time_message", "")
        # 任务中心顺序执行模式：回调转交任务中心处理
        if self._task_mode:
            def _tc_finish():
                self.task_center.on_finish(return_code, elapsed, task)
            self.root.after(0, _tc_finish)
            return

        def _apply():
            if return_code == 0:
                status = STATUS_SUCCESS
                detail = ""
                self._log(f"\n[成功] {task.name} 执行完成")
                self.logger.info(f"执行成功: {task.name}")
                self._set_status(f"完成: {task.name} (用时 {elapsed:.1f}s)")
            else:
                status = STATUS_FAILED
                detail = login_message or trading_time_message or f"退出码: {return_code}"
                self._log(f"\n[错误] {task.name} 执行失败，退出码: {return_code}")
                self.logger.error(f"执行失败: {task.name}, 退出码: {return_code}")
                self._set_status(f"失败: {task.name} (用时 {elapsed:.1f}s)")

            if self._current_record_id is not None:
                self.history.update_record(self._current_record_id, status, elapsed, detail)
                self._current_record_id = None
                self._refresh_history()

            if task.category == "交易系统设置" and self._single_settings_batch:
                task_status = (
                    TaskCenter.ST_STOPPED
                    if self._single_stop_requested
                    else TaskCenter.ST_SUCCESS if return_code == 0
                    else TaskCenter.ST_FAILED
                )
                self.update_settings_batch(
                    self._single_settings_batch,
                    [{
                        "category": "交易系统设置",
                        "script_name": task.name,
                        "script_path": task.path,
                        "params": task.params,
                        "status": task_status,
                        "return_code": return_code,
                        "elapsed": elapsed,
                        "error": detail,
                    }],
                    final=True,
                    stopped=self._single_stop_requested,
                )
                self._single_settings_batch = None
                self._single_stop_requested = False

            self._reset_running_state()
        self.root.after(0, _apply)

    def _on_run_error(self, exc, task):
        # 任务中心顺序执行模式：回调转交任务中心处理
        if self._task_mode:
            def _tc_error():
                self.task_center.on_error(exc, task)
            self.root.after(0, _tc_error)
            return

        def _apply():
            self._log(f"\n[异常] 执行出错: {exc}")
            self.logger.error(f"执行异常: {exc}")
            self._set_status(f"异常: {task.name}")

            if self._current_record_id is not None:
                self.history.update_record(self._current_record_id, STATUS_ERROR, detail=str(exc))
                self._current_record_id = None
                self._refresh_history()

            if task.category == "交易系统设置" and self._single_settings_batch:
                self.update_settings_batch(
                    self._single_settings_batch,
                    [{
                        "category": "交易系统设置",
                        "script_name": task.name,
                        "script_path": task.path,
                        "params": task.params,
                        "status": (
                            TaskCenter.ST_STOPPED
                            if self._single_stop_requested
                            else TaskCenter.ST_ERROR
                        ),
                        "return_code": None,
                        "elapsed": 0.0,
                        "error": str(exc),
                    }],
                    final=True,
                    stopped=self._single_stop_requested,
                )
                self._single_settings_batch = None
                self._single_stop_requested = False

            self._reset_running_state()
        self.root.after(0, _apply)

    def _reset_running_state(self):
        """执行结束后复位运行状态与按钮"""
        self.is_running = False
        self.execute_btn.config(state=tk.NORMAL)
        self.stop_btn.config(state=tk.DISABLED)

    # ====================== 状态栏 ======================
    def _set_status(self, text, running=False):
        """设置状态栏（线程安全，可在子线程调用）"""
        self.root.after(0, self._apply_status, text, running)

    def _apply_status(self, text, running):
        """在主线程更新状态栏"""
        self.status_label.config(text=text)
        if running:
            self._status_running = True
            self.task_start_time = time.time()
            self._tick_timer()
        else:
            self._status_running = False
            self.time_label.config(text="")

    def _tick_timer(self):
        """定时刷新运行时间"""
        if not self._status_running:
            return
        elapsed = time.time() - self.task_start_time
        self.time_label.config(text=f"运行时间: {elapsed:.1f}s")
        self.root.after(200, self._tick_timer)

    def _on_log_level_change(self):
        """切换日志级别并持久化"""
        self.user_config["log_level"] = self.log_level.get()
        save_user_config(self.user_config)
        self.level_label.config(text=f"日志: {self.log_level.get()}")
        self._log(f"日志级别已切换为: {self.log_level.get()}")
        self.logger.info(f"日志级别切换为: {self.log_level.get()}")

    def _show_log_level_help(self):
        """日志级别说明"""
        messagebox.showinfo(
            "日志级别说明",
            "详细日志：显示窗口/GUI 操作信息以及子脚本的 print 输出。\n\n"
            "简洁日志：仅显示窗口/GUI 的关键信息（开始、完成、错误等），"
            "不显示子脚本内部的 print 输出，界面更清爽，适合普通用户。\n\n"
            "当前级别可通过「设置」菜单随时切换，默认「详细日志」。"
        )

    def _log(self, message):
        """输出日志（带颜色区分，线程安全）"""
        self.root.after(0, self.log_text.append, message)
