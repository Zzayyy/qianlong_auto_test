# -*- coding: utf-8 -*-
"""
钱龙期权交易 - GUI自动化工具 v2
===========================
功能:
    集成所有查询、下单、组合申报自动化脚本
    提供图形界面一键执行，带参数配置和日志输出

运行环境:
    pip install pywinauto openpyxl pandas

使用方法:
    1. 打开钱龙旗舰版，登录交易账号
    2. 运行本GUI工具
    3. 通过菜单选择功能分类，选择脚本，配置参数后执行
"""

import os
import sys
import time
import json
import threading
import subprocess
import logging
import re
from datetime import datetime
from pathlib import Path
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog

# ====================== PyInstaller 打包兼容 ======================
# 判断是否从 PyInstaller 打包的 exe 运行
IS_FROZEN = getattr(sys, 'frozen', False)

if IS_FROZEN:
    # 打包后：脚本在 _internal 目录下
    BASE_DIR = os.path.join(os.path.dirname(sys.executable), '_internal')
else:
    # 开发环境：脚本所在目录
    BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ====================== 脚本运行器模式 ======================
# 打包后exe自身充当Python解释器，用于在无Python环境的电脑上执行子脚本
# 用法: main.exe --_run_script <脚本路径>
def _run_script_mode():
    """
    脚本运行器模式：
    当exe收到 --_run_script 参数时，不启动GUI，而是直接执行指定的Python脚本。
    这样子脚本就能复用exe自带的Python环境和所有打包进去的依赖包。
    """
    if '--_run_script' not in sys.argv:
        return

    # 强制设置 stdout/stderr 编码为 UTF-8，避免特殊字符编码错误
    # 同时设置 line_buffering=True，确保 print 输出实时传递到父进程管道
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace', line_buffering=True)
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace', line_buffering=True)

    # 打包环境下，统一添加 pywin32 DLL 搜索路径
    # 这样所有子脚本（包括不导入 core.window 的脚本）都能正常加载 pywinauto
    if IS_FROZEN:
        dll_dir = os.path.join(os.path.dirname(sys.executable), "_internal", "pywin32_system32")
        if os.path.exists(dll_dir):
            os.environ['PATH'] = dll_dir + os.pathsep + os.environ.get('PATH', '')
            try:
                os.add_dll_directory(dll_dir)
            except AttributeError:
                pass

    idx = sys.argv.index('--_run_script')
    if idx + 1 >= len(sys.argv):
        print("错误: --_run_script 后需要指定脚本路径")
        sys.exit(1)

    script_path = sys.argv[idx + 1]

    if not os.path.exists(script_path):
        print(f"错误: 脚本文件不存在: {script_path}")
        sys.exit(1)

    # 设置子脚本环境
    sys.argv = sys.argv[idx + 1:]  # 子脚本看到的 argv[0] 是脚本路径

    # 确保子脚本所在目录在 sys.path 中
    script_dir = os.path.dirname(os.path.abspath(script_path))
    if script_dir not in sys.path:
        sys.path.insert(0, script_dir)

    # 确保 _internal 目录也在 sys.path 中（打包后依赖包在这里）
    if IS_FROZEN and BASE_DIR not in sys.path:
        sys.path.insert(0, BASE_DIR)

    try:
        import runpy
        runpy.run_path(script_path, run_name='__main__')
    except SystemExit:
        raise
    except Exception as e:
        print(f"执行脚本失败: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

    sys.exit(0)


# ====================== 用户配置 ======================
# 配置文件路径：打包后放在exe同级目录，开发环境放在脚本同级目录
if IS_FROZEN:
    CONFIG_DIR = os.path.dirname(sys.executable)
else:
    CONFIG_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(CONFIG_DIR, "config.json")
DEFAULT_OUTPUT_DIR = r"E:\Code\3\output"
CATEGORIES = ["查询", "下单", "组合申报"]


def load_user_config() -> dict:
    """加载用户配置"""
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                cfg = json.load(f)
                # 统一路径斜杠为 Windows 反斜杠
                if "output_dirs" in cfg:
                    for cat in cfg["output_dirs"]:
                        cfg["output_dirs"][cat] = cfg["output_dirs"][cat].replace("/", "\\")
                return cfg
        except Exception:
            pass
    return {
        "output_dirs": {cat: DEFAULT_OUTPUT_DIR for cat in CATEGORIES},
        "auto_open": False,
        "export_format": "txt"
    }


def save_user_config(config: dict):
    """保存用户配置"""
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"保存配置失败: {e}")


def get_output_dir(config: dict, category: str) -> str:
    """获取指定界面的输出目录"""
    dirs = config.get("output_dirs", {})
    return dirs.get(category, DEFAULT_OUTPUT_DIR)


def set_output_dir(config: dict, category: str, path: str):
    """设置指定界面的输出目录"""
    if "output_dirs" not in config:
        config["output_dirs"] = {}
    config["output_dirs"][category] = path.replace("/", "\\")


def get_script_filename(script_name: str) -> str:
    """从脚本名提取纯文件名（去掉编号和空格）"""
    # "3. 策略委托" -> "策略委托"
    name = re.sub(r"^\d+\.\s*", "", script_name).strip()
    return name


# ====================== 脚本配置 ======================
# 项目根目录（打包后指向 _internal，开发环境指向项目根目录）
PROJECT_ROOT = BASE_DIR

SCRIPTS_CONFIG = {
    "查询": [
        {"name": "1. 资金持仓", "path": rf"{PROJECT_ROOT}\查询\1.资金持仓.py"},
        {"name": "2. 策略持仓", "path": rf"{PROJECT_ROOT}\查询\2.策略持仓.py"},
        {"name": "3. 策略委托", "path": rf"{PROJECT_ROOT}\查询\3.策略委托.py"},
        {"name": "4. 历史策略持仓", "path": rf"{PROJECT_ROOT}\查询\4.历史策略持仓.py"},
        {"name": "5. 当日委托", "path": rf"{PROJECT_ROOT}\查询\5.当日委托.py"},
        {"name": "6. 当日成交", "path": rf"{PROJECT_ROOT}\查询\6.当日成交.py"},
        {"name": "7. 历史委托", "path": rf"{PROJECT_ROOT}\查询\7.历史委托.py"},
        {"name": "8. 历史成交", "path": rf"{PROJECT_ROOT}\查询\8.历史成交.py"},
        {"name": "9. 期权合约", "path": rf"{PROJECT_ROOT}\查询\9.期权合约.py"},
        {"name": "10. 资金查询", "path": rf"{PROJECT_ROOT}\查询\10.资金查询.py"},
        {"name": "11. 当日行权被指派查询", "path": rf"{PROJECT_ROOT}\查询\11.当日行权被指派查询.py"},
        {"name": "12. 历史结果查询", "path": rf"{PROJECT_ROOT}\查询\12.历史结果查询.py"},
        {"name": "13. 历史行权指派查询", "path": rf"{PROJECT_ROOT}\查询\13.历史行权指派查询.py"},
        {"name": "14. 历史对账单查询", "path": rf"{PROJECT_ROOT}\查询\14.历史对账单查询.py"},
        {"name": "15. 历史交割单查询", "path": rf"{PROJECT_ROOT}\查询\15.历史交割单查询.py"},
        {"name": "16. 历史行权交割查询", "path": rf"{PROJECT_ROOT}\查询\16.历史行权交割查询.py"},
        {"name": "17. 客户限仓信息查询", "path": rf"{PROJECT_ROOT}\查询\17.客户限仓信息查询.py"},
        {"name": "18. 限购额度查询", "path": rf"{PROJECT_ROOT}\查询\18.限购额度查询.py"},
        {"name": "19. 行权负债信息查询", "path": rf"{PROJECT_ROOT}\查询\19.行权负债信息查询.py"},
        {"name": "20. 历史行权负债信息", "path": rf"{PROJECT_ROOT}\查询\20.历史行权负债信息.py"},
    ],
    "下单": [
        {"name": "1.期权下单_自动化下单", "path": rf"{PROJECT_ROOT}\下单\自动化下单\4.期权下单(新)_自动化下单_Excel驱动版.py"},
        #{"name": "2.期权持仓_平仓/反手自动化", "path": rf"{PROJECT_ROOT}\下单\表格\9.Excel驱动_OCR定位_平仓操作.py"},
        {"name": "2.三键下单_自动化下单", "path": rf"{PROJECT_ROOT}\下单\自动化下单\4.三键下单_自动化下单_Excel驱动版.py"},
        {"name": "3.四键下单_自动化下单", "path": rf"{PROJECT_ROOT}\下单\自动化下单\4.四键下单_自动化下单_Excel驱动版.py"},
        {"name": "4.期权持仓_平仓/反手自动化_RapidOCR", "path": rf"{PROJECT_ROOT}\下单\表格\10.Excel驱动_OCR_RapidOCR.py"},
        {"name": "5.期权下单_一键导出", "path": rf"{PROJECT_ROOT}\下单\自动化导出\期权下单(新)_自动导出.py"},
        {"name": "6.全选撤单", "path": rf"{PROJECT_ROOT}\撤单\撤单_全选撤单_自动化.py"},
    ],
    "组合申报": [
        {"name": "1.组合申报_全自动", "path": rf"{PROJECT_ROOT}\组合申报\2.组合申报_全自动.py"},
        {"name": "2.拆分申报_全自动", "path": rf"{PROJECT_ROOT}\组合申报\2.拆分申报_全自动.py"},
        {"name": "3.组合策略持仓查询", "path": rf"{PROJECT_ROOT}\组合申报\查询\1.组合策略持仓查询.py"},
        {"name": "4.组合策略信息查询", "path": rf"{PROJECT_ROOT}\组合申报\查询\2.组合策略信息查询.py"},
        {"name": "5.组合委托流水查询", "path": rf"{PROJECT_ROOT}\组合申报\查询\3.组合委托流水查询.py"},
        {"name": "6.历史组合委托流水", "path": rf"{PROJECT_ROOT}\组合申报\查询\4.历史组合委托流水.py"},
    ],
}


class AutomationGUI:
    """GUI自动化主界面"""

    def __init__(self, root):
        self.root = root
        self.root.title("钱龙期权交易 - GUI自动化工具")
        self.root.geometry("950x750")
        self.root.minsize(850, 650)

        self.current_process = None
        self.is_running = False
        self.current_category = "查询"

        # 加载用户配置
        self.user_config = load_user_config()

        # 参数变量（使用配置中的默认值）
        self.export_format = tk.StringVar(value=self.user_config.get("export_format", "txt"))
        self.auto_open = tk.BooleanVar(value=self.user_config.get("auto_open", False))
        self.txt_path = tk.StringVar(value="")
        self.xls_path = tk.StringVar(value="")
        self.order_qty = tk.IntVar(value=1)
        self.countdown_sec = tk.IntVar(value=3)
        self.xlsx_file = tk.StringVar(value="")

        # 期权下单_一键导出 参数
        self.export_target_position = tk.BooleanVar(value=True)  # 持仓
        self.export_target_order = tk.BooleanVar(value=True)     # 委托
        self.export_output_dir = tk.StringVar(value=get_output_dir(self.user_config, "下单"))

        # 日志目录：打包后放在exe同级目录
        if IS_FROZEN:
            self.log_dir = os.path.join(os.path.dirname(sys.executable), "logs")
        else:
            self.log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
        os.makedirs(self.log_dir, exist_ok=True)

        self._setup_logging()
        self._build_ui()
        self.logger.info("GUI自动化工具启动")

    def _setup_log_tags(self):
        """配置日志颜色标签"""
        self.log_text.tag_configure("success", foreground="#4ec9b0")
        self.log_text.tag_configure("error", foreground="#f44747")
        self.log_text.tag_configure("warning", foreground="#dcdcaa")
        self.log_text.tag_configure("info", foreground="#569cd6")
        self.log_text.tag_configure("separator", foreground="#808080")
        self.log_text.tag_configure("highlight", foreground="#ce9178")

    def _get_log_tag(self, message):
        """根据日志内容动态返回颜色标签"""
        msg_lower = message.lower()
        msg_stripped = message.strip()

        # 成功类（绿色）
        if any(kw in msg_lower for kw in ["成功", "完成", "success", "[ok]", "已找到", "已切换", "已点击", "已设置", "已选择", "主动打开"]):
            return "success"
        # 错误类（红色）
        elif any(kw in msg_lower for kw in ["错误", "失败", "error", "异常", "exception", "未找到", "找不到", "超时"]):
            return "error"
        # 警告类（黄色）
        elif any(kw in msg_lower for kw in ["警告", "warn", "提示", "注意", "[停止]", "请", "倒计时", "秒后"]):
            return "warning"
        # 配置信息类（蓝色）
        elif any(kw in msg_lower for kw in ["开始执行", "切换分类", "导出格式", "自动打开", "路径", "excel文件", "委托数量", "目标:", "窗口", "输出路径", "脚本路径", "txt路径", "xls路径"]):
            return "info"
        # 分隔线（灰色）
        elif msg_stripped.startswith("=") or msg_stripped.startswith("-") * 10:
            return "separator"
        # 高亮数据（橙色）
        elif any(kw in msg_lower for kw in ["[预览]", "字段", "共", "行"]):
            return "highlight"
        return None

    def _setup_logging(self):
        """配置日志系统"""
        self.logger = logging.getLogger("AutomationGUI")
        self.logger.setLevel(logging.DEBUG)
        log_file = os.path.join(
            self.log_dir,
            f"gui_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
        )
        file_handler = logging.FileHandler(log_file, encoding='utf-8')
        file_handler.setLevel(logging.DEBUG)
        fmt = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(fmt)
        self.logger.addHandler(file_handler)

    def _build_ui(self):
        """构建界面"""
        # 菜单栏
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # 功能菜单
        func_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="功能", menu=func_menu)
        func_menu.add_command(label="下单", command=lambda: self._switch_category("下单"))
        func_menu.add_command(label="组合申报", command=lambda: self._switch_category("组合申报"))
        func_menu.add_command(label="查询", command=lambda: self._switch_category("查询"))

        # 工具菜单
        tool_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="工具", menu=tool_menu)
        tool_menu.add_command(label="清空日志", command=self._clear_log)
        tool_menu.add_command(label="打开日志目录", command=self._open_log_dir)
        tool_menu.add_separator()
        tool_menu.add_command(label="退出", command=self.root.quit)

        # 主框架
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)

        # 标题
        ttk.Label(
            main_frame,
            text="钱龙期权交易自动化",
            font=("Microsoft YaHei UI", 16, "bold")
        ).pack(pady=(0, 10))

        # 左右分栏（PanedWindow 保证日志区域始终可见）
        self.paned = ttk.PanedWindow(main_frame, orient=tk.HORIZONTAL)
        self.paned.pack(fill=tk.BOTH, expand=True)

        # 左侧面板：脚本列表 + 参数配置
        left_frame = ttk.Frame(self.paned)
        self.paned.add(left_frame, weight=1)

        # 操作按钮 - 固定在底部
        self.btn_frame = ttk.Frame(left_frame)
        self.btn_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=(5, 0))

        self.execute_btn = ttk.Button(self.btn_frame, text="执行", command=self._execute_script, width=12)
        self.execute_btn.pack(side=tk.LEFT, padx=5)

        self.stop_btn = ttk.Button(self.btn_frame, text="停止", command=self._stop_script, state=tk.DISABLED, width=10)
        self.stop_btn.pack(side=tk.LEFT, padx=5)

        # 当前分类标签
        self.category_label = ttk.Label(
            left_frame,
            text="",
            font=("Microsoft YaHei UI", 12, "bold"),
            foreground="#0078d4"
        )
        self.category_label.pack(side=tk.TOP, anchor=tk.W, pady=(0, 5))

        # 中间内容区域 - 占据剩余空间
        middle_frame = ttk.Frame(left_frame)
        middle_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, pady=(0, 10))

        # 脚本列表
        list_frame = ttk.LabelFrame(middle_frame, text="脚本列表", padding="8")
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 10))

        self.script_listbox = tk.Listbox(
            list_frame,
            font=("Microsoft YaHei UI", 10),
            selectmode=tk.SINGLE,
            exportselection=False,
            height=12
        )
        self.script_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar = ttk.Scrollbar(list_frame, orient=tk.VERTICAL, command=self.script_listbox.yview)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.script_listbox.config(yscrollcommand=scrollbar.set)
        self.script_listbox.bind('<Double-Button-1>', lambda e: self._execute_script())
        self.script_listbox.bind('<<ListboxSelect>>', lambda e: (self._update_paths_for_selected_script(), self._update_params_for_selected_script()))

        # 参数配置面板
        self.params_frame = ttk.LabelFrame(middle_frame, text="参数配置", padding="8")
        self.params_frame.pack(fill=tk.X, pady=(0, 10))

        # 数据预览面板（仅下单显示）
        self.preview_frame = ttk.LabelFrame(middle_frame, text="数据预览", padding="3")
        # 初始隐藏

        # 预览Treeview + 滚动条
        tree_wrap = ttk.Frame(self.preview_frame)
        tree_wrap.pack(fill=tk.X, anchor=tk.W)

        self.preview_tree = ttk.Treeview(tree_wrap, show="headings", height=5)
        v_scroll = ttk.Scrollbar(tree_wrap, orient=tk.VERTICAL, command=self.preview_tree.yview)
        h_scroll = ttk.Scrollbar(self.preview_frame, orient=tk.HORIZONTAL, command=self.preview_tree.xview)
        self.preview_tree.configure(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)

        self.preview_tree.pack(side=tk.LEFT, fill=tk.BOTH)
        v_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        h_scroll.pack(side=tk.BOTTOM, fill=tk.X)

        # 预览信息标签
        self.preview_info = ttk.Label(self.preview_frame, text="选择Excel文件后显示数据预览", foreground="gray")
        self.preview_info.pack(anchor=tk.W, pady=(3, 0))

        # 右侧：日志
        log_frame = ttk.LabelFrame(self.paned, text="运行日志", padding="5")
        self.paned.add(log_frame, weight=1)

        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            font=("Consolas", 9),
            state=tk.DISABLED,
            wrap=tk.WORD,
            bg="#1e1e1e",
            fg="#d4d4d4",
            insertbackground="#d4d4d4",
            width=50
        )
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # 配置日志颜色标签
        self._setup_log_tags()

        # 延迟设置分栏比例，确保窗口渲染完成
        self.root.after(200, lambda: self._init_sash_position(main_frame))

        # 默认显示查询
        self._switch_category("下单")

    def _init_sash_position(self, main_frame):
        """初始化分栏位置"""
        try:
            total_width = main_frame.winfo_width()
            if total_width > 0:
                self.paned.sashpos(0, int(total_width * 0.55))
        except Exception:
            pass

    def _reset_sash(self):
        """重置分栏位置"""
        try:
            total_width = self.paned.winfo_width()
            if total_width > 0:
                self.paned.sashpos(0, int(total_width * 0.55))
        except Exception:
            pass

    def _switch_category(self, category):
        """切换分类"""
        self.current_category = category
        self.category_label.config(text=f"当前功能: {category}")
        self.script_listbox.delete(0, tk.END)

        scripts = SCRIPTS_CONFIG.get(category, [])
        for script in scripts:
            self.script_listbox.insert(tk.END, script["name"])

        # 重建参数面板
        self._rebuild_params()

        # 仅下单显示数据预览
        if self.current_category == "下单":
            self.preview_frame.pack(fill=tk.X, pady=(0, 10))
        else:
            self.preview_frame.pack_forget()

        self._log(f"切换分类: {category}")
        self.logger.info(f"切换分类: {category}")

    def _update_paths_for_selected_script(self):
        """根据当前选中的脚本更新路径显示"""
        script = self._get_selected_script()
        if not script:
            return
        output_dir = get_output_dir(self.user_config, self.current_category)
        filename = get_script_filename(script["name"])
        self.txt_path.set(os.path.join(output_dir, f"{filename}.txt"))
        self.xls_path.set(os.path.join(output_dir, f"{filename}.xls"))

    def _rebuild_params(self):
        """根据分类重建参数配置面板"""
        for w in self.params_frame.winfo_children():
            w.destroy()

        if self.current_category == "查询":
            self._build_query_params()
        elif self.current_category == "下单":
            self._build_order_params()
        elif self.current_category == "组合申报":
            self._build_combo_params()

    def _update_params_for_selected_script(self):
        """根据选中的脚本更新参数面板"""
        if self.current_category != "下单":
            return
        script = self._get_selected_script()
        if not script:
            return

        # 先清空参数配置面板
        for w in self.params_frame.winfo_children():
            w.destroy()

        # 根据脚本决定是否显示数据预览
        if script["name"] in ("5.期权下单_一键导出", "6.全选撤单"):
            # 这两个脚本不需要数据预览
            self.preview_frame.pack_forget()
        else:
            # 其他脚本需要数据预览
            self.preview_frame.pack(fill=tk.X, pady=(0, 10))

        # 根据脚本更新参数配置面板
        if script["name"] == "6.全选撤单":
            # 全选撤单：显示提示信息
            ttk.Label(
                self.params_frame,
                text="该功能无需参数配置，点击执行即可。",
                foreground="gray"
            ).pack(pady=20)
        elif script["name"] == "5.期权下单_一键导出":
            # 期权下单_一键导出：显示参数配置
            self._build_export_params()
        else:
            # 其他脚本：显示订单参数配置
            self._build_order_params()

    def _build_query_params(self):
        """查询类参数"""
        ttk.Label(self.params_frame, text="导出格式:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(self.params_frame, text="TXT", variable=self.export_format, value="txt").grid(row=0, column=1, sticky=tk.W)
        ttk.Radiobutton(self.params_frame, text="XLS", variable=self.export_format, value="xls").grid(row=0, column=2, sticky=tk.W)

        ttk.Label(self.params_frame, text="自动打开文件:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Checkbutton(self.params_frame, text="导出后自动打开", variable=self.auto_open).grid(row=1, column=1, sticky=tk.W, columnspan=2)

        ttk.Label(self.params_frame, text="TXT输出路径:").grid(row=2, column=0, sticky=tk.W, pady=5)
        txt_entry = ttk.Entry(self.params_frame, textvariable=self.txt_path, width=35)
        txt_entry.grid(row=2, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=lambda: self._browse_path(self.txt_path, ".txt")).grid(row=2, column=2, padx=5)

        ttk.Label(self.params_frame, text="XLS输出路径:").grid(row=3, column=0, sticky=tk.W, pady=5)
        xls_entry = ttk.Entry(self.params_frame, textvariable=self.xls_path, width=35)
        xls_entry.grid(row=3, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=lambda: self._browse_path(self.xls_path, ".xls")).grid(row=3, column=2, padx=5)

        self.params_frame.columnconfigure(1, weight=1)

    def _build_order_params(self):
        """下单参数"""
        # 参数行
        param_row = ttk.Frame(self.params_frame)
        param_row.pack(fill=tk.X)

        ttk.Label(param_row, text="Excel文件:").grid(row=0, column=0, sticky=tk.W, pady=5)
        xlsx_entry = ttk.Entry(param_row, textvariable=self.xlsx_file, width=35)
        xlsx_entry.grid(row=0, column=1, sticky=tk.EW, pady=5)
        ttk.Button(param_row, text="选择文件", command=self._browse_xlsx).grid(row=0, column=2, padx=5)
        param_row.columnconfigure(1, weight=1)

        self.params_frame.columnconfigure(0, weight=1)

    def _build_combo_params(self):
        """组合申报参数"""
        ttk.Label(self.params_frame, text="导出格式:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Radiobutton(self.params_frame, text="TXT", variable=self.export_format, value="txt").grid(row=0, column=1, sticky=tk.W)
        ttk.Radiobutton(self.params_frame, text="XLS", variable=self.export_format, value="xls").grid(row=0, column=2, sticky=tk.W)

        ttk.Label(self.params_frame, text="TXT输出路径:").grid(row=1, column=0, sticky=tk.W, pady=5)
        txt_entry = ttk.Entry(self.params_frame, textvariable=self.txt_path, width=35)
        txt_entry.grid(row=1, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=lambda: self._browse_path(self.txt_path, ".txt")).grid(row=1, column=2, padx=5)

        ttk.Label(self.params_frame, text="XLS输出路径:").grid(row=2, column=0, sticky=tk.W, pady=5)
        xls_entry = ttk.Entry(self.params_frame, textvariable=self.xls_path, width=35)
        xls_entry.grid(row=2, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=lambda: self._browse_path(self.xls_path, ".xls")).grid(row=2, column=2, padx=5)

        ttk.Label(self.params_frame, text="自动打开文件:").grid(row=3, column=0, sticky=tk.W, pady=5)
        ttk.Checkbutton(self.params_frame, text="导出后自动打开", variable=self.auto_open).grid(row=3, column=1, sticky=tk.W, columnspan=2)

        ttk.Separator(self.params_frame, orient=tk.HORIZONTAL).grid(row=4, column=0, columnspan=3, sticky=tk.EW, pady=10)

        ttk.Label(self.params_frame, text="委托数量:").grid(row=5, column=0, sticky=tk.W, pady=5)
        ttk.Spinbox(self.params_frame, from_=1, to=999, textvariable=self.order_qty, width=10).grid(row=5, column=1, sticky=tk.W, pady=5)

        self.params_frame.columnconfigure(1, weight=1)

    def _build_export_params(self):
        """期权下单_一键导出 参数配置"""
        # 清空旧内容
        for widget in self.params_frame.winfo_children():
            widget.destroy()

        # 导出目标选择
        ttk.Label(self.params_frame, text="导出目标:").grid(row=0, column=0, sticky=tk.W, pady=5)
        ttk.Checkbutton(self.params_frame, text="持仓", variable=self.export_target_position).grid(row=0, column=1, sticky=tk.W, padx=5)
        ttk.Checkbutton(self.params_frame, text="委托", variable=self.export_target_order).grid(row=0, column=2, sticky=tk.W, padx=5)

        # 输出目录配置
        ttk.Label(self.params_frame, text="输出目录:").grid(row=1, column=0, sticky=tk.W, pady=5)
        path_entry = ttk.Entry(self.params_frame, textvariable=self.export_output_dir, width=35)
        path_entry.grid(row=1, column=1, sticky=tk.EW, pady=5)
        ttk.Button(self.params_frame, text="浏览", command=self._browse_export_dir).grid(row=1, column=2, padx=5)

        # 显示文件名格式说明
        ttk.Label(self.params_frame, text="文件名格式: 期权下单(新)-持仓-20260629.xls", foreground="gray").grid(row=2, column=0, columnspan=3, sticky=tk.W, pady=5)

        self.params_frame.columnconfigure(1, weight=1)

        # 确保按钮框架可见
        self.btn_frame.lift()

    def _browse_path(self, var, ext):
        """浏览选择文件路径"""
        initial_dir = get_output_dir(self.user_config, self.current_category)
        # 默认文件名：当前输入框已有值，或根据选中脚本生成
        current = var.get()
        if current and os.path.basename(current):
            initial_file = os.path.basename(current)
        else:
            script = self._get_selected_script()
            if script:
                initial_file = get_script_filename(script["name"]) + ext
            else:
                initial_file = "output" + ext

        path = filedialog.asksaveasfilename(
            initialdir=initial_dir,
            initialfile=initial_file,
            defaultextension=ext,
            filetypes=[("All files", "*.*"), (f"{ext.upper()} files", f"*{ext}")]
        )
        if path:
            var.set(path)
            new_dir = os.path.dirname(path)
            if new_dir != get_output_dir(self.user_config, self.current_category):
                set_output_dir(self.user_config, self.current_category, new_dir)
                save_user_config(self.user_config)
                self._log(f"[配置] 已更新{self.current_category}输出目录: {new_dir}")

    def _browse_xlsx(self):
        """选择Excel文件"""
        path = filedialog.askopenfilename(
            title="选择Excel配置文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self.xlsx_file.set(path)
            self._log(f"已选择Excel文件: {path}")
            self._preview_excel(path)

    def _browse_export_dir(self):
        """选择导出目录"""
        path = filedialog.askdirectory(
            title="选择导出目录",
            initialdir=self.export_output_dir.get()
        )
        if path:
            self.export_output_dir.set(path)
            self._log(f"已设置导出目录: {path}")

    def _preview_excel(self, filepath: str):
        """读取Excel并显示到预览表格"""
        if not HAS_OPENPYXL:
            self.preview_info.config(text="未安装 openpyxl，无法预览。请运行: pip install openpyxl", foreground="red")
            return

        # 清空旧数据
        self.preview_tree.delete(*self.preview_tree.get_children())
        self.preview_tree["columns"] = ()

        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active

            # 使用 iter_rows 安全读取，避免直接访问 ws[1] 导致索引错误
            rows_iter = ws.iter_rows(values_only=True)

            # 读取第一行作为表头
            try:
                first_row = next(rows_iter)
            except StopIteration:
                self.preview_info.config(text="Excel文件无任何数据", foreground="red")
                wb.close()
                return

            if not first_row or all(v is None for v in first_row):
                self.preview_info.config(text="Excel文件无表头或为空", foreground="red")
                wb.close()
                return

            headers = [str(v) if v is not None else "" for v in first_row]

            if all(h == "" for h in headers):
                self.preview_info.config(text="Excel文件无表头或为空", foreground="red")
                wb.close()
                return

            # 设置列（紧凑宽度，不拉伸）
            self.preview_tree["columns"] = headers
            for h in headers:
                self.preview_tree.heading(h, text=h)
                self.preview_tree.column(h, width=60, minwidth=30, stretch=False)

            # 读取数据行（最多显示50行）
            row_count = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or all(v is None for v in row):
                    continue
                vals = [str(v) if v is not None else "" for v in row]
                # 补齐列数
                while len(vals) < len(headers):
                    vals.append("")
                self.preview_tree.insert("", tk.END, values=vals[:len(headers)])
                row_count += 1
                if row_count >= 50:
                    break

            total_rows = sum(1 for r in ws.iter_rows(min_row=2, values_only=True) if r and any(v is not None for v in r))
            show_note = f"（共{total_rows}行，仅显示前50行）" if total_rows > 50 else f"（共{total_rows}行）"
            self.preview_info.config(text=f"字段: {len(headers)} 个 | 数据: {show_note}", foreground="green")

            wb.close()

            # 数据加载后重置分栏位置，防止挤掉日志
            self.root.after(50, self._reset_sash)

        except Exception as e:
            self.preview_info.config(text=f"读取失败: {e}", foreground="red")
            self._log(f"[预览] Excel读取失败: {e}")

    def _get_selected_script(self):
        """获取选中的脚本"""
        selection = self.script_listbox.curselection()
        if not selection:
            return None
        script_name = self.script_listbox.get(selection[0])
        for scripts in SCRIPTS_CONFIG.values():
            for s in scripts:
                if s["name"] == script_name:
                    return s
        return None

    def _execute_script(self):
        """执行脚本"""
        if self.is_running:
            messagebox.showwarning("提示", "有脚本正在运行中，请先停止")
            return

        script = self._get_selected_script()
        if not script:
            messagebox.showwarning("提示", "请先选择一个脚本")
            return

        if not os.path.exists(script["path"]):
            messagebox.showerror("错误", f"脚本文件不存在:\n{script['path']}")
            return

        # 下单需要检查Excel文件（全选撤单和期权下单_一键导出除外）
        if self.current_category == "下单" and script["name"] not in ("5.期权下单_一键导出", "6.全选撤单") and not self.xlsx_file.get():
            messagebox.showwarning("提示", "请先选择Excel配置文件")
            return

        # 期权下单_一键导出需要检查导出目标
        if script["name"] == "5.期权下单_一键导出":
            if not self.export_target_position.get() and not self.export_target_order.get():
                messagebox.showwarning("提示", "请至少选择一个导出目标（持仓或委托）")
                return

        # 保存当前配置
        self.user_config["export_format"] = self.export_format.get()
        self.user_config["auto_open"] = self.auto_open.get()
        save_user_config(self.user_config)

        # 确保路径已设置
        if self.current_category == "查询" and not self.txt_path.get():
            self._update_paths_for_selected_script()

        self.is_running = True
        self.execute_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)

        self._log(f"\n{'='*60}")
        self._log(f"开始执行: {script['name']}")
        self._log(f"脚本路径: {script['path']}")

        # 打印参数
        if self.current_category == "查询":
            self._log(f"导出格式: {self.export_format.get().upper()}")
            self._log(f"自动打开: {'是' if self.auto_open.get() else '否'}")
            self._log(f"TXT路径: {self.txt_path.get()}")
            self._log(f"XLS路径: {self.xls_path.get()}")
        elif self.current_category == "下单":
            if script["name"] == "5.期权下单_一键导出":
                export_targets = []
                if self.export_target_position.get():
                    export_targets.append("持仓")
                if self.export_target_order.get():
                    export_targets.append("委托")
                self._log(f"导出目标: {', '.join(export_targets)}")
                self._log(f"输出目录: {self.export_output_dir.get()}")
                self._log(f"文件名格式: 期权下单(新)-持仓-20260629.xls")
            else:
                self._log(f"Excel文件: {self.xlsx_file.get()}")
        elif self.current_category == "组合申报":
            self._log(f"导出格式: {self.export_format.get().upper()}")
            self._log(f"自动打开: {'是' if self.auto_open.get() else '否'}")
            self._log(f"TXT路径: {self.txt_path.get()}")
            self._log(f"XLS路径: {self.xls_path.get()}")

        self._log(f"{'='*60}")
        self.logger.info(f"开始执行: {script['name']}")

        thread = threading.Thread(
            target=self._run_script,
            args=(script,),
            daemon=True
        )
        thread.start()

    def _run_script(self, script):
        """运行脚本"""
        try:
            # 构建环境变量，将GUI参数传递给子脚本
            env = os.environ.copy()
            env["PYTHONPATH"] = PROJECT_ROOT  # 让子脚本能找到 core 模块
            env["GUI_EXPORT_FORMAT"] = self.export_format.get()
            env["GUI_AUTO_OPEN"] = str(self.auto_open.get())
            env["GUI_TXT_PATH"] = self.txt_path.get()
            env["GUI_XLS_PATH"] = self.xls_path.get()
            env["GUI_ORDER_QTY"] = str(self.order_qty.get())
            env["GUI_COUNTDOWN"] = str(self.countdown_sec.get())
            env["GUI_XLSX_FILE"] = self.xlsx_file.get()
            env["GUI_CATEGORY"] = self.current_category

            # 期权下单_一键导出 参数
            export_targets = []
            if self.export_target_position.get():
                export_targets.append("持仓")
            if self.export_target_order.get():
                export_targets.append("委托")
            env["GUI_EXPORT_TARGETS"] = ",".join(export_targets)
            env["GUI_EXPORT_DIR"] = self.export_output_dir.get()

            # 构建命令：打包后用exe自身充当Python解释器，开发环境用系统Python
            if IS_FROZEN:
                cmd = [sys.executable, "--_run_script", script["path"]]
            else:
                cmd = [sys.executable, "-u", script["path"]]

            self._log(f"目标: {cmd[0]}")
            self.logger.info(f"执行命令: {cmd}")

            env["PYTHONIOENCODING"] = "utf-8"  # 子进程用UTF-8输出，避免gbk编码错误
            env["PYTHONUTF8"] = "1"

            self.current_process = subprocess.Popen(
                cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding='utf-8',
                errors='replace',
                creationflags=subprocess.CREATE_NO_WINDOW,
                env=env
            )

            for line in self.current_process.stdout:
                line = line.rstrip()
                if line:
                    self._log(line)
                    self.logger.debug(line)

            return_code = self.current_process.wait()

            if return_code == 0:
                self._log(f"\n[成功] {script['name']} 执行完成")
                self.logger.info(f"执行成功: {script['name']}")
            else:
                self._log(f"\n[错误] {script['name']} 执行失败，退出码: {return_code}")
                self.logger.error(f"执行失败: {script['name']}, 退出码: {return_code}")

        except Exception as e:
            self._log(f"\n[异常] 执行出错: {e}")
            self.logger.error(f"执行异常: {e}")

        finally:
            self.is_running = False
            self.current_process = None
            self.execute_btn.config(state=tk.NORMAL)
            self.stop_btn.config(state=tk.DISABLED)

    def _stop_script(self):
        """停止脚本"""
        if self.current_process and self.current_process.poll() is None:
            self._log("\n[停止] 用户手动停止...")
            self.logger.info("用户手动停止")
            try:
                self.current_process.terminate()
                self.current_process.wait(timeout=3)
            except Exception:
                try:
                    self.current_process.kill()
                except Exception:
                    pass

    def _clear_log(self):
        """清空日志"""
        self.log_text.config(state=tk.NORMAL)
        self.log_text.delete(1.0, tk.END)
        self.log_text.config(state=tk.DISABLED)

    def _open_log_dir(self):
        """打开日志目录"""
        os.startfile(self.log_dir)

    def _log(self, message):
        """输出日志（带颜色区分）"""
        def _append():
            self.log_text.config(state=tk.NORMAL)
            ts = datetime.now().strftime("%H:%M:%S")
            tag = self._get_log_tag(message)
            if tag:
                self.log_text.insert(tk.END, f"[{ts}] {message}\n", tag)
            else:
                self.log_text.insert(tk.END, f"[{ts}] {message}\n")
            self.log_text.see(tk.END)
            self.log_text.config(state=tk.DISABLED)
        self.root.after(0, _append)


def main():
    root = tk.Tk()
    app = AutomationGUI(root)
    root.mainloop()


if __name__ == "__main__":
    # 脚本运行器模式：如果收到 --_run_script 参数，执行子脚本后直接退出
    _run_script_mode()
    main()
