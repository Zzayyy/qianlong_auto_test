# -*- coding: utf-8 -*-
import sys
import os
_here = os.path.dirname(os.path.abspath(__file__))
for _d in (_here, os.path.dirname(_here), os.path.dirname(os.path.dirname(_here)), os.path.dirname(os.path.dirname(os.path.dirname(_here)))):
    if os.path.isdir(os.path.join(_d, "core")) and _d not in sys.path:
        sys.path.insert(0, _d)
        break
from core.window import find_window, activate_window, switch_panel as switch_main_panel

"""
钱龙期权交易 - 三键下单面板自动化(Excel文件驱动版)
==============================================
功能流程:
    1. 从 Excel 文件读取下单配置(菜单/合约代码/报价方式/委托数量/动作/备兑/自动/FOK)
    2. 查找"钱龙"主窗口
    3. 激活窗口并切换到"三键下单"面板
    4. 按 Excel 每行配置执行下单:
       - 录入合约代码
       - 选择报价方式
       - 设置委托数量
       - 勾选备兑/自动/FOK
       - 点击对应动作按钮(买开/卖开/平仓)
    5. 确认弹窗

Excel 字段说明:
    菜单     : 面板名称(如"三键下单")
    合约代码 : 期权合约代码
    报价方式 : 对手价/挂盘价/涨停价/跌停价/限价/超价/市价转限/市价FAK/市价FOK
    委托数量 : 纯数字(如"5") / 百分比(如"30%") / "FOK"
    动作     : 买开 / 卖开 / 平仓
    备兑     : True=勾选, False=不勾选
    自动     : True=勾选, False=不勾选
    FOK      : True=勾选, False=不勾选

运行环境:
    pip install pywinauto==0.6.8 openpyxl

使用方法:
    1. 打开钱龙旗舰版,登录交易账号
    2. 准备 Excel 文件,填入下单配置(菜单列填写"三键下单")
    3. 修改下方 EXCEL_PATH 为你的 Excel 文件路径
    4. 运行本脚本(将鼠标快速移到屏幕左上角可紧急停止)
    5. 倒计时内切换到钱龙窗口
"""

from pywinauto import Application, findwindows
import os
import time
import sys
import ctypes
import openpyxl
import win32gui
import win32con
import win32api
import win32process


# ====================== 可配置参数 ======================
# ---- Excel 路径:优先读取 GUI 传入的环境变量,否则使用默认路径 ----
_gui_xlsx = os.environ.get("GUI_XLSX_FILE", "").strip()
if _gui_xlsx and os.path.exists(_gui_xlsx):
    EXCEL_PATH = _gui_xlsx
else:
    EXCEL_PATH = r"C:\Users\Administrator\Desktop\新建 XLSX 工作表.xlsx"
WINDOW_KEY = "钱龙模拟期权宝"        # 窗口标题关键字
USE_ENTER_CONFIRM = True  # True=按回车(推荐), False=鼠标点"确定"
COUNTDOWN = 3              # 操作前倒计时秒数
INTERVAL = 1.0             # 每两次下单之间的间隔(秒)
# ========================================================


# 报价方式映射: Excel中填写值 -> 实际下拉框显示值
QUOTE_TYPE_MAP = {
    "市价FOK": "市价F0K",   # 实际控件是 F0K(数字0),不是 FOK(字母O)
}

# 报价方式下拉项(以软件实际显示为准;顺序与快速下单页 owner-drawn ComboLBox 实机核对一致)
QUOTE_TYPES = (
    "对手价",
    "挂盘价",
    "涨停价",
    "跌停价",
    "限价",
    "超价",
    "市价转限",
    "市价FAK",
    "市价F0K",
)
QUOTE_INDEX = {name: idx for idx, name in enumerate(QUOTE_TYPES)}

QUOTE_AUTO_ID = "18083"       # 报价方式输入框 auto_id
QTY_AUTO_ID   = "306"         # 委托数量输入框 auto_id

# 下单按钮映射: 动作名称 -> 控件 auto_id
ACTION_BUTTONS = {
    "买开": "18066",   # 实际 title: 买多
    "卖开": "18065",   # 实际 title: 卖空
    "平仓": "18070",   # 实际 title: 平仓
}

# 复选框 auto_id 映射
CHECKBOX_AUTO_IDS = {
    "备兑": "5025",
    "FOK": "5027",
    "自动": "5028",
}


def read_excel(path: str) -> list:
    """读取 Excel 文件,返回配置列表(只读取菜单列为"三键下单"的行)。"""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    # 标准化列名(去除空格)
    headers = [str(h).strip() if h else "" for h in headers]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        item = dict(zip(headers, row))
        # 只读取菜单列为"三键下单"的行
        menu_value = str(item.get("菜单", "")).strip()
        if menu_value != "三键下单":
            continue
        rows.append(item)

    wb.close()
    print(f"[OK] 读取 Excel: {path}, 共 {len(rows)} 条配置(菜单列过滤: 三键下单)")
    return rows





def switch_panel(win, tree_item: str):
    # 先让 TreeView 滚到顶部,无论滚轮当前在哪
    tree = win.child_window(auto_id="1223", control_type="Tree")
    tree.wait("ready", timeout=10)
    tree.set_focus()
    tree.type_keys("{HOME}", with_spaces=False)
    time.sleep(0.2)
    # 再点击目标节点
    item = win.child_window(title=tree_item, control_type="TreeItem")  # HOME后默认是期权下单，不然有的电脑可有问题，蓝色选中的就识别不到
    #item.wait("visible", timeout=10)
    # item.click_input()
    item.select()
    print(f"[OK] 已切换到面板: {tree_item}")


# ====================== win32gui 加速辅助 ======================
def _find_child_by_ctrl_id(parent_hwnd, ctrl_id):
    """枚举 parent_hwnd 的所有子窗口,返回控件 ID 匹配 ctrl_id 的句柄列表(快于 UIA 查找)。"""
    found = []
    def _cb(hwnd, _):
        try:
            if win32gui.GetDlgCtrlID(hwnd) == ctrl_id:
                found.append(hwnd)
        except Exception:
            pass
    try:
        win32gui.EnumChildWindows(parent_hwnd, _cb, None)
    except Exception:
        pass
    return found


def _first_visible(hwnds):
    """从候选句柄中取第一个可见的;都为不可见时返回第一个;空则返回 None。"""
    if not hwnds:
        return None
    return next((h for h in hwnds if win32gui.IsWindowVisible(h)), hwnds[0])


def _mouse_click(hwnd):
    """用真实鼠标点击控件中心(可靠触发软件点击逻辑,等价于 click_input 但无 UIA 查找开销)。"""
    # 尽量把所属顶层窗口置前,确保点击落在正确位置
    try:
        root = win32gui.GetAncestor(hwnd, 2)  # GA_ROOT = 2
        if not root:
            root = win32gui.GetParent(hwnd)
        if root:
            win32gui.SetForegroundWindow(root)
    except Exception:
        pass
    time.sleep(0.02)
    l, t, r, b = win32gui.GetWindowRect(hwnd)
    x = (l + r) // 2
    y = (t + b) // 2
    win32api.SetCursorPos((x, y))
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0)
    win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0)


def _find_windows_by_text(main_hwnd, text):
    """查找文字等于 text 的可见窗口(先查主窗口子窗口,再查同进程顶层弹窗)。

    用于定位下拉项(ListItem/Button):这些项在 UIA 中是 ListItem,
    在 win32 中是带文字的独立窗口。
    """
    found = []
    def _child_cb(hwnd, _):
        try:
            if win32gui.GetWindowText(hwnd) == text and win32gui.IsWindowVisible(hwnd):
                found.append(hwnd)
        except Exception:
            pass
    try:
        win32gui.EnumChildWindows(main_hwnd, _child_cb, None)
    except Exception:
        pass
    if found:
        return found
    # 主窗口子窗口未命中,再查同进程顶层弹窗(如下拉列表)
    try:
        _, pid = win32process.GetWindowThreadProcessId(main_hwnd)
    except Exception:
        pid = None
    def _top_cb(hwnd, _):
        try:
            if win32gui.GetWindowText(hwnd) == text and win32gui.IsWindowVisible(hwnd):
                if pid is not None:
                    _, wpid = win32process.GetWindowThreadProcessId(hwnd)
                    if wpid != pid:
                        return
                found.append(hwnd)
        except Exception:
            pass
    try:
        win32gui.EnumWindows(_top_cb, None)
    except Exception:
        pass
    return found


def _edit_set_text(edit_hwnd, text):
    """直接对 Edit 控件发送 WM_SETTEXT(等价于 pywinauto 的 set_edit_text,但无 UIA 查找开销)。"""
    try:
        win32gui.SetFocus(edit_hwnd)
    except Exception:
        pass
    win32gui.SendMessage(edit_hwnd, win32con.WM_SETTEXT, 0, str(text))


def _press_enter(edit_hwnd):
    """仅向编辑框发送回车(兜底用)。"""
    try:
        win32gui.SetForegroundWindow(win32gui.GetParent(edit_hwnd) or edit_hwnd)
    except Exception:
        pass
    try:
        win32gui.SetFocus(edit_hwnd)
    except Exception:
        pass
    time.sleep(0.1)
    win32api.keybd_event(0x0D, 0, 0, 0)
    win32api.keybd_event(0x0D, 0, win32con.KEYEVENTF_KEYUP, 0)


def fill_contract_code(main_hwnd, code: str):
    """录入合约代码(直接 WM_SETTEXT 整段替换,无鼠标/UIA 查找开销)。"""
    edit = _first_visible(_find_child_by_ctrl_id(main_hwnd, 18005))
    if not edit:
        print("[WARN] 未找到合约代码输入框(auto_id=18005),跳过")
        return
    _edit_set_text(edit, code)
    print(f"[OK] 已录入合约代码: {code}")


def click_lock(main_hwnd):
    """点击"锁定"按钮,触发合约信息刷新(win32gui 直接 BM_CLICK)。"""
    btn = _first_visible(_find_windows_by_text(main_hwnd, "锁定"))
    if not btn:
        print("[WARN] 未找到'锁定'按钮,跳过")
        return
    win32gui.SendMessage(btn, win32con.BM_CLICK, 0, 0)
    print("[OK] 已点击'锁定'按钮")


def confirm_auto_dialog(main_win=None, timeout: float = 8):
    """确认'自动净仓'弹窗(回车确认)。

    通过进程 ID 过滤,避免把回车发到其它程序窗口
    (例如打开着的 "平仓.xlsx - Excel")。
    """
    main_hwnd = None
    main_pid = None
    if main_win is not None:
        try:
            main_hwnd = main_win.handle
        except Exception:
            main_hwnd = None
        try:
            main_pid = main_win.process_id()
        except Exception:
            if main_hwnd is not None:
                pid = ctypes.c_ulong(0)
                ctypes.windll.user32.GetWindowThreadProcessId(
                    main_hwnd, ctypes.byref(pid))
                main_pid = pid.value or None

    end = time.time() + timeout
    while time.time() < end:
        try:
            elems = findwindows.find_elements(top_level_only=True)
        except Exception:
            elems = []
        for elem in elems:
            # 进程过滤: 只处理与主窗口同进程的窗口
            if main_pid is not None:
                try:
                    if elem.process_id != main_pid:
                        continue
                except Exception:
                    continue
            hwnd = elem.handle
            if main_hwnd is not None and hwnd == main_hwnd:
                continue
            try:
                title = (elem.name or "").strip()
            except Exception:
                title = ""
            if "自动净仓" not in title:
                continue
            try:
                dlg_app = Application(backend="uia").connect(handle=hwnd, timeout=0.5)
                dlg = dlg_app.window(handle=hwnd)
                dlg.set_focus()
                time.sleep(0.1)
                dlg.type_keys("{ENTER}", with_spaces=False)
                print(f"[OK] 自动净仓弹窗已确认 (hwnd={hwnd}, title='{title}', pid={main_pid})")
                return True
            except Exception:
                continue
        time.sleep(0.15)
    print(f"[WARN] 等待'自动净仓'弹窗超时({timeout}s)")


def get_checkbox(main_hwnd, name: str, cache: dict = None):
    """获取复选框控件句柄,支持缓存(同一会话内复用,避免重复查找)。"""
    if name not in CHECKBOX_AUTO_IDS:
        raise ValueError(f"未知复选框: {name!r}")
    ctrl_id = int(CHECKBOX_AUTO_IDS[name])
    if cache is not None:
        cached = cache.get(name)
        if cached is not None and win32gui.IsWindow(cached):
            return cached
        elif cached is not None:
            cache.pop(name, None)
    matches = _find_child_by_ctrl_id(main_hwnd, ctrl_id)
    h = _first_visible(matches)
    if h is None:
        # 控件可能因其它选项(如勾选"自动"会让"备兑"置灰)而暂时不可见/不可用
        return None
    if cache is not None:
        cache[name] = h
    return h


def set_checkbox(main_hwnd, name: str, enable: bool, cache: dict = None, main_win=None):
    """设置复选框到目标状态(用真实鼠标点击切换,确保触发软件逻辑)。

    优化点:
      1. 用 win32gui 按控件 ID 定位句柄(快于 UIA 查找)
      2. 仅在当前状态与目标不一致时才点击,无冗余动作
      3. 鼠标点击失效时回退到 BM_CLICK 消息,保证可靠性
    """
    if name not in CHECKBOX_AUTO_IDS:
        raise ValueError(f"未知复选框: {name!r}")

    cb = get_checkbox(main_hwnd, name, cache)
    if cb is None:
        print(f"[WARN] 复选框 {name} 未找到(可能受其它选项影响被置灰/隐藏),跳过")
        return

    # 检查是否为灰色(不可用)状态
    if not win32gui.IsWindowEnabled(cb):
        if enable:
            print(f"[WARN] 复选框 {name} 为灰色(不可用),无法勾选(可能'自动'已勾选),跳过")
        else:
            print(f"[WARN] 复选框 {name} 为灰色(不可用),跳过")
        return

    is_checked = win32gui.SendMessage(cb, win32con.BM_GETCHECK, 0, 0) != 0
    if enable == is_checked:
        print(f"[WARN] 复选框 {name} 已是{'勾选' if enable else '取消勾选'}状态,跳过")
        return

    # 用真实鼠标点击切换(等价于 click_input,确保触发软件逻辑,如"自动"会弹出自动净仓弹窗)
    try:
        _mouse_click(cb)
    except Exception:
        try:
            win32gui.SendMessage(cb, win32con.BM_CLICK, 0, 0)
        except Exception:
            pass
    time.sleep(0.08)

    # 验证是否生效,未生效则回退到 BM_CLICK
    try:
        if (win32gui.SendMessage(cb, win32con.BM_GETCHECK, 0, 0) != 0) != enable:
            try:
                win32gui.SendMessage(cb, win32con.BM_CLICK, 0, 0)
            except Exception:
                _mouse_click(cb)
    except Exception:
        pass

    if enable:
        print(f"[OK] 勾选复选框: {name}")
        # 勾选"自动"时会弹出"自动净仓"确认弹窗,需回车确认
        if name == "自动":
            time.sleep(0.2)
            confirm_auto_dialog(main_win=main_win)
    else:
        print(f"[OK] 取消勾选复选框: {name}")


def set_all_checkboxes(main_hwnd, enable_beidui: bool, enable_zidong: bool, enable_fok: bool, cache: dict = None, main_win=None):
    """统一设置备兑、自动、FOK 三个复选框(直接设定到目标状态,无需预重置)。

    设置顺序说明:必须先确定"自动"的最终状态,再设"备兑"。
    因为软件中勾选"自动"后"备兑"会置灰,若先设"备兑"再设"自动",
    上一单遗留的"自动=勾选"会让本单"备兑"一开始就是灰色而设不上去。
    先设"自动":若目标为取消,则"备兑"解冻;若目标为勾选,则"备兑"本就该保持未勾选。
    """
    set_checkbox(main_hwnd, "自动", enable_zidong, cache=cache, main_win=main_win)
    set_checkbox(main_hwnd, "备兑", enable_beidui, cache=cache, main_win=main_win)
    set_checkbox(main_hwnd, "FOK", enable_fok, cache=cache, main_win=main_win)


def _press_vk(vk_code: int) -> None:
    """发送一次虚拟键按下/释放(不依赖窗口焦点在控件上的 UIA 路径)。"""
    win32api.keybd_event(vk_code, 0, 0, 0)
    win32api.keybd_event(vk_code, 0, win32con.KEYEVENTF_KEYUP, 0)


def _quote_text_matches(edit_hwnd, expected: str) -> bool:
    """校验报价方式输入框文本是否为目标值。

    钱龙自绘 Edit 跨进程 GetWindowText/UIA 可能拿不到文本,
    用 WM_GETTEXTLENGTH 比对: 字符数与 GBK 字节数任一匹配即可
    (兼容 Unicode 窗口返回字符数、ANSI/GBK 窗口返回字节数两种情况)。
    """
    try:
        actual = win32gui.GetWindowText(edit_hwnd) or ""
        if actual:
            return actual.strip() == expected
    except Exception:
        pass
    try:
        n = win32gui.SendMessage(edit_hwnd, win32con.WM_GETTEXTLENGTH, 0, 0)
        return n == len(expected) or n == len(expected.encode("gbk"))
    except Exception:
        return False


def _quote_verified(edit_hwnd, expected: str, index=None) -> bool:
    """报价方式是否已选中。

    实测(2026-07-31 中泰真机): 该自绘控件已选中"挂盘价"时 CB_GETCURSEL 仍返回 0,
    索引不可用作判定依据; GetWindowText 也恒为空。唯一可靠的是 WM_GETTEXTLENGTH:
    写入目标文本后长度等于目标(字符数或 GBK 字节数)即视为生效。
    """
    return _quote_text_matches(edit_hwnd, expected)


def select_quote_type(main_hwnd, option: str, auto_id: str = QUOTE_AUTO_ID):
    """选择报价方式。

    该控件是钱龙自绘的"可输入下拉"(UIA 暴露为 Edit,右缘带隐藏下拉箭头),
    点击输入框中心只会聚焦编辑区、不会展开下拉(owner-drawn 下拉项也不是
    带文字的独立窗口,不能按 GetWindowText 定位)。因此:

      1. 优先 WM_SETTEXT 直接写入目标文本再回车(可编辑下拉按文本匹配,不弹下拉;
         实测 2026-07-31 中泰真机文本写入即生效);
      2. 仅当文本写入未生效时,才点右缘箭头展开下拉,按 HOME + N×DOWN + ENTER 选择
         (与快速下单页在钱龙实机验证过的 owner-drawn ComboBox 方案一致);
      3. 校验只用 WM_GETTEXTLENGTH 文本长度(字符数或 GBK 字节数),
         不用 CB_GETCURSEL(实测该自绘控件索引不可信,已选中仍返回 0)。
      注: 不设"当前已是目标"短路——同长度文本(对手价/挂盘价/涨停价都是 3 字)
      无法靠长度区分,直接重写目标文本最稳妥。
    """
    actual_option = QUOTE_TYPE_MAP.get(option, option)
    index = QUOTE_INDEX.get(actual_option)

    edit = _first_visible(_find_child_by_ctrl_id(main_hwnd, int(auto_id)))
    if not edit:
        print(f"[WARN] 未找到报价方式输入框(auto_id={auto_id}),跳过")
        return

    # 1) 直接写文本 + 回车(利用可编辑下拉的文本匹配,不展开下拉)
    _edit_set_text(edit, actual_option)
    time.sleep(0.15)
    _press_enter(edit)
    time.sleep(0.25)
    if _quote_verified(edit, actual_option, index):
        print(f"[OK] 报价方式: {option} -> {actual_option} (文本写入+回车)")
        return

    # 2) 点右缘箭头展开下拉 + 键盘选择(仅当前一步失败才执行,正常只会展开一次)
    if index is not None:
        try:
            win32gui.SetForegroundWindow(win32gui.GetParent(edit) or edit)
        except Exception:
            pass
        l, t, r, b = win32gui.GetWindowRect(edit)
        cy = (t + b) // 2
        for px in (r - 12, r - 4):
            win32api.SetCursorPos((px, cy))
            time.sleep(0.05)
            win32api.mouse_event(win32con.MOUSEEVENTF_LEFTDOWN, 0, 0)
            win32api.mouse_event(win32con.MOUSEEVENTF_LEFTUP, 0, 0)
            time.sleep(0.3)
            _press_vk(win32con.VK_HOME)
            for _ in range(index):
                _press_vk(win32con.VK_DOWN)
            _press_vk(win32con.VK_RETURN)
            time.sleep(0.25)
            if _quote_verified(edit, actual_option, index):
                print(f"[OK] 报价方式: {option} -> {actual_option} (下拉键盘选择)")
                return

    print(f"[WARN] 报价方式未校验生效: {option} -> {actual_option},可能仍为默认值")


def click_action_button(main_hwnd, action: str):
    """点击下单动作按钮(买开/卖开/平仓),用真实鼠标点击(确保触发软件下单逻辑)。"""
    if action not in ACTION_BUTTONS:
        raise ValueError(f"未知动作: {action!r},可选: {list(ACTION_BUTTONS.keys())}")

    ctrl_id = int(ACTION_BUTTONS[action])
    btn = _first_visible(_find_child_by_ctrl_id(main_hwnd, ctrl_id))
    if not btn:
        print(f"[WARN] 未找到动作按钮: {action} (auto_id={ctrl_id}),跳过")
        return
    # 用真实鼠标点击(等价于 pywinauto click_input,但省去 UIA 查找开销),确保下单逻辑被触发
    try:
        _mouse_click(btn)
    except Exception:
        try:
            win32gui.SendMessage(btn, win32con.BM_CLICK, 0, 0)
        except Exception:
            pass
    print(f"[OK] 下单动作: {action} (auto_id={ctrl_id})")


def press_enter_to_confirm(
    main_win=None,
    dialog_patterns: list = None,
    timeout: float = 3,
):
    """通过回车键确认弹窗。

    通过进程 ID 过滤,避免把回车发送到其它程序窗口
    (例如打开着的 "平仓.xlsx - Excel" 也会命中 "平仓" 关键字)。
    """
    if dialog_patterns is None:
        dialog_patterns = ["三键下单", "提示", "确认", "确定"]

    main_hwnd = None
    main_pid = None
    if main_win is not None:
        try:
            main_hwnd = main_win.handle
        except Exception:
            main_hwnd = None
        try:
            main_pid = main_win.process_id()
        except Exception:
            if main_hwnd is not None:
                pid = ctypes.c_ulong(0)
                ctypes.windll.user32.GetWindowThreadProcessId(
                    main_hwnd, ctypes.byref(pid))
                main_pid = pid.value or None

    # 高优先级关键字(明确是弹窗)优先于宽泛关键字(平仓/反手)
    high_priority = ("提示", "确认", "确定", "委托确认", "风险提示")

    end = time.time() + timeout
    while time.time() < end:
        try:
            elems = findwindows.find_elements(top_level_only=True)
        except Exception:
            elems = []

        high_candidates = []
        low_candidates = []
        for elem in elems:
            # 1) 进程过滤: 只处理与主窗口同进程的窗口
            if main_pid is not None:
                try:
                    if elem.process_id != main_pid:
                        continue
                except Exception:
                    continue
            hwnd = elem.handle
            # 2) 跳过主窗口本身
            if main_hwnd is not None and hwnd == main_hwnd:
                continue
            try:
                title = (elem.name or "").strip()
            except Exception:
                title = ""
            if not any(p in title for p in dialog_patterns):
                continue
            if any(p in title for p in high_priority):
                high_candidates.append((hwnd, title))
            else:
                low_candidates.append((hwnd, title))

        for hwnd, title in high_candidates + low_candidates:
            try:
                dlg_app = Application(backend="uia").connect(handle=hwnd, timeout=0.5)
                dlg = dlg_app.window(handle=hwnd)
                dlg.set_focus()
                time.sleep(0.1)
                dlg.type_keys("{ENTER}", with_spaces=False)
                print(f"[OK] 回车确认 (hwnd={hwnd}, title='{title}', pid={main_pid})")
                return True
            except Exception as e:
                print(f"[--] 弹窗(hwnd={hwnd}, title='{title}')回车失败: {e}")
                continue

        time.sleep(0.15)
    print(f"[WARN] 等待弹窗({timeout}s),匹配: {dialog_patterns}")
    return False


def confirm_all_dialogs(
    main_win=None,
    max_dialogs: int = 5,
    no_dialog_timeout: float = 2.0,
    per_dialog_timeout: float = 3.0,
    use_enter: bool = USE_ENTER_CONFIRM,
):
    """自动确认所有弹窗，直到一段时间内没有新弹窗出现。
    
    Args:
        main_win: 主窗口对象,用于进程过滤,避免误操作其它程序窗口
        max_dialogs: 最大弹窗数量上限（防止死循环）
        no_dialog_timeout: 等待新弹窗的超时时间（秒），超过此时间无新弹窗则认为全部处理完毕
        per_dialog_timeout: 单个弹窗的等待超时时间（秒）
        use_enter: 是否使用回车确认
    """
    count = 0
    for i in range(1, max_dialogs + 1):
        print(f"[WARN] 等待第 {i} 个弹窗 ({no_dialog_timeout}s无新弹窗则结束)...")
        ok = press_enter_to_confirm(main_win=main_win, timeout=per_dialog_timeout)
        if ok:
            count += 1
            print(f"[OK] 已确认第 {count} 个弹窗")
            time.sleep(0.4)  # 弹窗间短暂间隔
        else:
            # 超时未出现弹窗，认为全部处理完毕
            print(f"[OK] 无更多弹窗，共确认 {count} 个")
            break
    else:
        print(f"[WARN] 达到最大弹窗数量上限({max_dialogs})")


def fill_order_quantity(main_hwnd, value, auto_id: str = QTY_AUTO_ID):
    """设置委托数量(数字直写 WM_SETTEXT;百分比/FOK 走下拉项点击)。"""
    str_value = str(value)

    # 纯数字路径
    if str_value.isdigit():
        edit = _first_visible(_find_child_by_ctrl_id(main_hwnd, int(auto_id)))
        if edit:
            _edit_set_text(edit, str_value)
            print(f"[OK] 委托数量(直写): {str_value}")
            return
        print(f"[WARN] 未找到委托数量输入框(auto_id={auto_id}),跳过")

    # 百分比 / FOK 路径(点击输入框弹出下拉,再点对应按钮)
    edit = _first_visible(_find_child_by_ctrl_id(main_hwnd, int(auto_id)))
    if not edit:
        print(f"[WARN] 未找到委托数量输入框(auto_id={auto_id}),跳过")
        return
    _mouse_click(edit)
    time.sleep(0.4)

    candidates = [str_value, str_value.replace("%", ""), f"{str_value}%"]
    for title in candidates:
        items = _find_windows_by_text(main_hwnd, title)
        if items:
            _mouse_click(items[0])
            print(f"[OK] 委托数量(下拉): {title}")
            return

    # 兜底:键盘输入
    print(f"[WARN] 下拉未匹配到 {str_value},改用键盘输入")
    _edit_set_text(edit, str_value)
    _press_enter(edit)


def countdown(seconds: int):
    """操作前倒计时,让用户有时间切换窗口。"""
    print(f"将在 {seconds} 秒后开始,请把焦点切到钱龙软件...")
    try:
        for i in range(seconds, 0, -1):
            print(f"  {i}...", end="\r")
            time.sleep(1)
    except KeyboardInterrupt:
        raise
    print(" " * 30, end="\r")


def normalize_bool(value) -> bool:
    """将 Excel 中的布尔值或字符串转换为 Python bool。"""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    if isinstance(value, str):
        s = value.strip().lower()
        if s in ("true", "是", "1", "yes"):
            return True
        elif s in ("false", "否", "0", "no"):
            return False
    return bool(value)


def main():
    try:
        # 读取 Excel 配置
        configs = read_excel(EXCEL_PATH)
        if not configs:
            print("[错误] Excel 文件为空或无有效数据")
            sys.exit(1)

        # 从第一行获取面板名称,未指定则用默认值
        first_config = configs[0]
        tree_item = first_config.get("菜单", "三键下单")
        if not tree_item:
            tree_item = "三键下单"

        print(f"计划执行: {len(configs)} 次下单")
        print(f"面板: {tree_item}")

        countdown(COUNTDOWN)

        hwnd = find_window(WINDOW_KEY)
        print(f"[OK] 已找到窗口,句柄 = {hwnd}")

        win = activate_window(hwnd)
        main_hwnd = win.handle  # 取原生句柄,后续操作全部走 win32gui,避免 pywinauto 查找开销
        switch_main_panel(win, tree_item)
        time.sleep(0.5)

        cb_cache = {}  # 复选框控件缓存,整个下单过程复用句柄

        for idx, cfg in enumerate(configs, 1):
            contract_code = str(cfg.get("合约代码", "")).strip()
            quote_type = str(cfg.get("报价方式", "")).strip()
            order_qty = cfg.get("委托数量", "1")
            # 小数 -> *100 + "%"
            if isinstance(order_qty, float):
                order_qty = f"{int(order_qty * 100)}%"
            else:
                order_qty = str(order_qty).strip()
            action = str(cfg.get("动作", "")).strip()
            enable_beidui = normalize_bool(cfg.get("备兑", False))
            enable_zidong = normalize_bool(cfg.get("自动", False))
            enable_fok = normalize_bool(cfg.get("FOK", False))

            print(f"\n=== [{idx}/{len(configs)}] 合约={contract_code} "
                  f"报价={quote_type} 动作={action} ===")

            if not contract_code:
                print("[WARN] 合约代码为空,跳过")
                continue

            # 录入合约代码
            fill_contract_code(main_hwnd, contract_code)
            time.sleep(0.3)

            # 设置委托数量
            fill_order_quantity(main_hwnd, order_qty)
            time.sleep(0.3)

            # 选择报价方式
            if quote_type:
                select_quote_type(main_hwnd, quote_type)
                time.sleep(0.3)

            # 直接设置复选框到目标状态(状态比较已保证正确,无需先全部取消)
            set_all_checkboxes(main_hwnd, enable_beidui, enable_zidong, enable_fok, cache=cb_cache, main_win=win)
            time.sleep(0.1)

            # 执行下单动作
            if action:
                click_action_button(main_hwnd, action)
                time.sleep(1)

                # 自动确认所有弹窗(无论2个还是3个)
                confirm_all_dialogs(main_win=win)

            time.sleep(INTERVAL)

        print(f"\n=== 全部完成: {len(configs)} 次下单 ===")

    except KeyboardInterrupt:
        print("\n[中断] 用户主动停止")
        sys.exit(0)
    except Exception as e:
        print(f"\n[错误] {type(e).__name__}: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
